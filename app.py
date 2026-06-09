"""文中 eHRS 排班/打卡 桌面工具(CustomTkinter)。

啟動:
    python3 app.py

第一次用請到「設定」分頁填帳號密碼並儲存(存本機 ~/.ehrs_tool,不進 git)。
排班分頁:選年月→載入→雙擊任一格可改班/刪班(會二次確認後才寫入正式班表)。
打卡分頁:選日期區間→查詢。
"""
import calendar as _cal
import datetime as dt
import re
import threading
import tkinter as tk
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache
from itertools import groupby
from tkinter import messagebox, ttk

import customtkinter as ctk

import tkinter.filedialog as filedialog

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from config import load_config, save_config
from ehrs_client import DEFAULT_BASE_URL, EhrsClient
from shift_codes_ref import SHIFT_CODES_REF
from version import __version__

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ─── UI 色彩字典（參考 HTML 設計稿配色）────────────────────────────────────────
_C = {
    "accent":  "#4A6FE3",   # 主色：靛藍
    "acc_h":   "#3B5DD0",   # 主色 hover
    "ink":     "#2D3048",   # 主文字：深藍灰
    "muted":   "#7B8097",   # 次文字：中灰
    "app_bg":  "#F5F7FB",   # 應用底色：極淡藍灰
    "tab_bg":  "#ECEEF5",   # 分頁列容器
    "line":    "#E8EBF4",   # 邊框 / 分隔線
    "white":   "#FFFFFF",
    "tbl_hdr": "#EEF2FB",   # 表格表頭
    "row_alt": "#F8F9FD",   # 隔行底色
    "red":     "#CC2200",
    "green":   "#2A8B5A",
}

_WEEK = "一二三四五六日"
_EXCEL_SHEET_BAD_CHARS = re.compile(r"[\[\]\*\?/\\:]")
_DEPT_KEYS = ("部門名稱", "部門", "部門全名", "pa51014FullName", "pa51014Name")
_DEPT_CODE_KEYS = ("部門代號", "pa51014")
_DEPT_FULL_KEYS = ("部門全名", "pa51014FullName")
_DEPT_NAME_KEYS = ("部門名稱", "部門", "pa51014Name")
_EMP_ID_KEYS = ("員工代號", "pa51002")
_EMP_NAME_KEYS = ("員工姓名", "pa51004")
_W_BTN = 86
_W_FILTER = 112
_W_PRIMARY = 124


def _excel_sheet_title(title: str) -> str:
    """Return a valid Excel worksheet title."""
    cleaned = _EXCEL_SHEET_BAD_CHARS.sub("-", title).strip().strip("'")
    return (cleaned or "Sheet")[:31]


def _first_field(row: dict, keys: tuple[str, ...]) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _get_schedule_periods(roc_year: int) -> list[tuple[dt.date, dt.date]]:
    """計算指定民國年度的 14 個四週排班期間（每期 28 天）。
    民國 115 年基準起始日：2025-12-22（週一）。
    相鄰年度以 ±52 週（364 天）近似推算，維持週一起始。
    回傳 [(start, end), ...] × 14。
    """
    _ANCHOR_ROC  = 115
    _ANCHOR_DATE = dt.date(2025, 12, 22)
    diff   = roc_year - _ANCHOR_ROC
    anchor = _ANCHOR_DATE + dt.timedelta(weeks=52 * diff)
    return [
        (anchor + dt.timedelta(days=28 * i),
         anchor + dt.timedelta(days=28 * i + 27))
        for i in range(14)
    ]


def _is_rest(code: str) -> bool:
    """排休班別判斷：991 及 991-x 開頭的都算排休。"""
    return bool(code) and code.startswith("991")


@lru_cache(maxsize=None)
def _shift_times(code: str) -> tuple[str, str]:
    """從 SHIFT_CODES_REF 解析班別的表定上班/下班時間（HH:MM）。
    SHIFT_CODES_REF 為唯讀常數,輸入班別代碼有限,故以 lru_cache 快取(熱路徑)。"""
    name = SHIFT_CODES_REF.get(code, ("", 3, 0.0))[0]
    m = re.search(r"(\d{2}:\d{2})-(\d{2}:\d{2})", name)
    return (m.group(1), m.group(2)) if m else ("", "")


@lru_cache(maxsize=None)
def _shift_hours(code: str) -> float:
    """從 SHIFT_CODES_REF 直接取得班別時數（已含休息扣除）。"""
    return SHIFT_CODES_REF.get(code, ("", 3, 0.0))[2]


def _punch_mins(t: str) -> int:
    h, m = t.split(":")
    return int(h) * 60 + int(m)


def _split_punches(punches: list[str], sched_in: str, sched_out: str) -> tuple[str, str]:
    """多筆打卡依距表定時間分為上班群/下班群，取上班群最早、下班群最晚。
    無排班資料時直接取第一筆和最後一筆。"""
    if not sched_in or not sched_out:
        return punches[0], punches[-1]
    si = _punch_mins(sched_in)
    so = _punch_mins(sched_out)
    if so < si:
        so += 1440
    in_grp, out_grp = [], []
    for p in punches:
        pm = _punch_mins(p)
        if pm < si - 120:
            pm += 1440
        (out_grp if abs(pm - so) <= abs(pm - si) else in_grp).append(p)
    return (in_grp[0] if in_grp else ""), (out_grp[-1] if out_grp else "")


def _split_punches_jiezhuan(punches: list[str], sched_in: str, sched_out: str) -> tuple[str, str]:
    """結轉打卡專用分群：不做 2 小時截斷，直接以距表定時間近遠判斷上/下班。
    支援跨夜班（凌晨打卡加 1440 修正）。"""
    if not sched_in or not sched_out:
        return punches[0], punches[-1]
    si = _punch_mins(sched_in)
    so = _punch_mins(sched_out)
    cross = so < si
    if cross:
        so += 1440   # 跨夜班：把下班時間加一天
    in_grp, out_grp = [], []
    for p in punches:
        pm = _punch_mins(p)
        # 跨夜班：凌晨打卡若距上班超過 6 小時，視為隔天
        if cross and pm < si - 360:
            pm += 1440
        (out_grp if abs(pm - so) <= abs(pm - si) else in_grp).append(p)
    return (in_grp[0] if in_grp else ""), (out_grp[-1] if out_grp else "")


def _overtime_hours(clock_in: str, clock_out: str, sched_in: str, sched_out: str,
                    kind: int | None = None) -> float:
    """加班 / 休息日加班時數(四捨五入至 0.5h);無加班回 0.0。
    為「備注字串」與「出勤統計」共用的單一計算來源(避免從字串反解)。"""
    # 休息日(kind=2):工作全程即加班
    if kind == 2:
        if not (clock_in and clock_out and sched_in and sched_out):
            return 0.0
        ci = _punch_mins(clock_in)
        co = _punch_mins(clock_out)
        if co < ci:
            co += 1440   # 跨夜
        worked = co - ci
        return round(worked / 60 * 2) / 2 if worked > 0 else 0.0
    # 一般工作日:下班超過表定 30 分才算加班(以打卡下班為準)
    if not (clock_out and sched_in and sched_out):
        return 0.0
    si = _punch_mins(sched_in)
    so = _punch_mins(sched_out)
    co = _punch_mins(clock_out)
    if so < si:
        so += 1440
    if co < si - 120:
        co += 1440
    if co > so + 30:
        return round((co - so) / 60 * 2) / 2
    return 0.0


def _punch_remark(clock_in: str, clock_out: str, sched_in: str, sched_out: str,
                  kind: int | None = None) -> str:
    """產生備注文字：遲到、早退、上班忘打卡、下班忘打卡、加班（可複合）。
    kind=2 休息日：有打卡即算「休息日加班 Xh」，不另做遲到/早退判斷。
    """
    remarks: list[str] = []

    # ── 休息日（kind=2）：只要有打卡就是加班 ──────────────────────
    if kind == 2:
        oh = _overtime_hours(clock_in, clock_out, sched_in, sched_out, kind=2)
        if oh > 0:
            remarks.append(f"休息日加班{oh:g}h")
        elif clock_in:
            remarks.append("休息日加班")
        return " ".join(remarks)

    # ── 一般工作日 ─────────────────────────────────────────────────
    if not clock_in:
        remarks.append("上班忘打卡")
    elif sched_in and _punch_mins(clock_in) > _punch_mins(sched_in):
        remarks.append("遲到")
    if not clock_out:
        remarks.append("下班忘打卡")
    elif sched_out and sched_in:
        si = _punch_mins(sched_in)
        so = _punch_mins(sched_out)
        co = _punch_mins(clock_out)
        if so < si:
            so += 1440
        if co < si - 120:
            co += 1440
        if co < so:
            remarks.append("早退")
        elif co > so + 30:                           # 超過 30 分鐘算加班
            oh = _overtime_hours(clock_in, clock_out, sched_in, sched_out, kind=kind)
            remarks.append(f"加班{oh:g}h")
    return " ".join(remarks)


# 備注分類所用關鍵字(集中一處,供 _classify_remark 與各分頁共用)
_REMARK_ERR_KEYWORDS = ("遲到", "上班忘打卡", "早退", "下班忘打卡", "曠職")
_REMARK_OT_KEYWORD = "加班"   # 含「加班」「休息日加班」


def _classify_remark(remark: str) -> str:
    """把備注歸成單一類別:'err'(異常) / 'ot'(加班) / 'leave'(假別) / ''(正常)。
    優先序 err > ot > leave。打卡分頁著色、篩選、統計、建議共用此判斷。"""
    remark = remark or ""
    if not remark:
        return ""
    if any(k in remark for k in _REMARK_ERR_KEYWORDS):
        return "err"
    if _REMARK_OT_KEYWORD in remark:
        return "ot"
    return "leave"


def _calc_hours(sched_in: str, sched_out: str) -> float:
    """計算實際工時（扣除休息）。
    ≤ 4h：不扣；> 4h 且 < 9h：扣 0.5h；≥ 9h：扣 1h。"""
    si = _punch_mins(sched_in)
    so = _punch_mins(sched_out)
    if so < si:
        so += 1440
    total = (so - si) / 60.0
    if total <= 4.0:
        return total
    return total - (1.0 if total >= 9.0 else 0.5)


def _suggest_shifts(
    clock_in: str, clock_out: str, shift_codes: dict,
    preferred_hours: float = 8.0, top_n: int = 3
) -> list[tuple]:
    """找最適班別：實際打卡時間不會造成遲到/早退的班別 top_n 名。
    排序規則：工時差距最小優先（同工時 → 時間最接近）。
    回傳 [(code, sched_in, sched_out, delta_in, delta_out, hours), ...]
    delta_in > 0 = 比排班提早到；delta_out > 0 = 比排班晚下班（加班）。
    """
    ci_m = _punch_mins(clock_in) if clock_in else None
    co_m = _punch_mins(clock_out) if clock_out else None
    if ci_m is None and co_m is None:
        return []
    candidates = []
    for code in shift_codes:
        if _is_rest(code):
            continue
        si, so = _shift_times(code)
        if not si or not so:
            continue
        si_m = _punch_mins(si)
        so_m = _punch_mins(so)
        if so_m < si_m:
            so_m += 1440
        ci_a = (ci_m + 1440) if ci_m is not None and ci_m < si_m - 240 else ci_m
        co_a = (co_m + 1440) if co_m is not None and co_m < si_m - 240 else co_m
        d_in  = (si_m - ci_a) if ci_a is not None else 0   # + = 早到
        d_out = (co_a - so_m) if co_a is not None else 0   # + = 加班
        if ci_a is not None and (d_in < 0 or d_in > 120):
            continue
        if co_a is not None and d_out < -5:
            continue
        h = _shift_hours(code) or _calc_hours(si, so)
        # 主要：工時差距（相差 0.5h = +1000 分）；次要：時間契合度
        score = abs(h - preferred_hours) * 1000 + d_in * 0.7 + max(0, d_out) * 0.5
        candidates.append((code, si, so, d_in, d_out, h, score))
    candidates.sort(key=lambda x: x[6])
    return [(c, si, so, d_in, d_out, h)
            for c, si, so, d_in, d_out, h, _ in candidates[:top_n]]


def _assign_single_punch(punch: str, sched_in: str, sched_out: str) -> tuple[str, str]:
    """只有一筆打卡時，依距表定上班/下班的近遠判斷是實際上班還是實際下班。
    回傳 (實際上班, 實際下班)，其中一個為空字串。"""
    if not sched_in and not sched_out:
        return punch, ""
    p = _punch_mins(punch)
    if sched_in and sched_out:
        si = _punch_mins(sched_in)
        so = _punch_mins(sched_out)
        if so < si:
            so += 1440
        if p < si - 120:
            p += 1440
        return ("", punch) if abs(p - so) <= abs(p - si) else (punch, "")
    return ("", punch) if sched_out else (punch, "")


_WHEEL_ACTIVE_TARGET = None
_WHEEL_ACTIVE_WIDGET = None
_WHEEL_BOUND = False


def _wheel_units(event) -> int:
    """把滾輪 / 觸控板事件換算成 yview_scroll 的 units（負值=往上）。
    跨平台：
      - macOS：delta 為小整數（±1、±2…），直接當單位數，不可除以 120/60。
      - Windows：delta 為 ±120 的倍數。
      - Linux/X11：無 delta，改用 Button-4（上）/ Button-5（下）。"""
    d = getattr(event, "delta", 0)
    if d:
        if abs(d) >= 120:                       # Windows / 部分 X11
            return int(-d / 120) or (-1 if d > 0 else 1)
        return -d                                # macOS：delta 已是單位數
    num = getattr(event, "num", 0)
    if num == 4:
        return -1
    if num == 5:
        return 1
    if num == 6:
        return -1
    if num == 7:
        return 1
    return 0


def _wheel_axis(event) -> str:
    """判斷滾輪事件要導向垂直或水平捲動。"""
    num = getattr(event, "num", 0)
    if num in (6, 7):
        return "x"
    # Tk 在多數平台把水平滾輪/觸控板事件映射成 Shift-MouseWheel。
    if getattr(event, "state", 0) & 0x0001:
        return "x"
    return "y"


def _scroll_target(target, event) -> bool:
    n = _wheel_units(event)
    if not n:
        return False
    axis = _wheel_axis(event)
    try:
        if axis == "x" and hasattr(target, "xview_scroll"):
            target.xview_scroll(n, "units")
        elif axis == "x" and hasattr(target, "xview"):
            target.xview("scroll", n, "units")
        elif hasattr(target, "yview_scroll"):
            target.yview_scroll(n, "units")
        else:
            target.yview("scroll", n, "units")
    except Exception:
        return False
    return True


def _bind_mousewheel(widget, target=None) -> None:
    """讓 widget 區域內的上下滾輪/觸控板捲動 target（預設為 widget 本身）。

    為何用全域 active target 而非每個 widget 都 bind_all / unbind_all：
    在 macOS + CustomTkinter 的巢狀版面中，滾輪事件不一定會送達被捲動的
    那個 widget（可能落到外層 CTk frame），導致只有少數頁面能捲。改採
    「全域只綁一次，指標進入時切換目前 target」的做法，避免離開某個表格時
    unbind_all 清掉 CustomTkinter 自己的觸控板捲動綁定。
    跨平台：macOS delta 為小整數、Windows 為 ±120 倍數、Linux 用 Button-4/5。
    也支援 Shift/水平滾輪導向 xview_scroll。"""
    global _WHEEL_BOUND
    tgt = target if target is not None else widget

    def _handler(event):
        if _scroll_target(tgt, event):
            return "break"
        return None

    def _global_handler(event):
        if _WHEEL_ACTIVE_TARGET is None:
            return None
        if _scroll_target(_WHEEL_ACTIVE_TARGET, event):
            return "break"
        return None

    def _on_enter(_e):
        global _WHEEL_ACTIVE_TARGET, _WHEEL_ACTIVE_WIDGET
        _WHEEL_ACTIVE_TARGET = tgt
        _WHEEL_ACTIVE_WIDGET = widget

    def _on_leave(_e):
        global _WHEEL_ACTIVE_TARGET, _WHEEL_ACTIVE_WIDGET
        if _WHEEL_ACTIVE_WIDGET is widget:
            _WHEEL_ACTIVE_TARGET = None
            _WHEEL_ACTIVE_WIDGET = None

    if not _WHEEL_BOUND:
        widget.bind_all("<MouseWheel>", _global_handler, add="+")
        widget.bind_all("<Shift-MouseWheel>", _global_handler, add="+")
        widget.bind_all("<Button-4>", _global_handler, add="+")
        widget.bind_all("<Button-5>", _global_handler, add="+")
        widget.bind_all("<Button-6>", _global_handler, add="+")
        widget.bind_all("<Button-7>", _global_handler, add="+")
        _WHEEL_BOUND = True

    # 直接綁在 widget 上作為後備（事件確實送達時即可用）
    widget.bind("<MouseWheel>", _handler)
    widget.bind("<Shift-MouseWheel>", _handler)
    widget.bind("<Button-4>", _handler)
    widget.bind("<Button-5>", _handler)
    widget.bind("<Button-6>", _handler)
    widget.bind("<Button-7>", _handler)
    # 指標進入/離開時切換 app 層級綁定，涵蓋事件未直接送達 widget 的情況
    widget.bind("<Enter>", _on_enter)
    widget.bind("<Leave>", _on_leave)


def _bind_scrollable_frame_mousewheel(scrollable) -> None:
    """補強 CTkScrollableFrame 的觸控板/滾輪事件，不覆蓋 CTk 內建綁定。"""
    target = (
        getattr(scrollable, "_parent_canvas", None)
        or getattr(scrollable, "_canvas", None)
        or scrollable
    )
    _bind_mousewheel(scrollable, target=target)


class _ScheduleHeader(tk.Canvas):
    """自訂雙排表頭：上列顯示月/日，下列顯示星期，六日紅字底色。
    與排班 Treeview 水平捲動同步（由外部呼叫 sync_x）。"""
    DATE_H  = 22
    WD_H    = 18
    TOTAL_H = DATE_H + WD_H   # 40 px
    EMP_W   = 150
    HOURS_W = 60    # 總時數欄
    COL_W   = 52
    _BG     = "#EEF2FB"
    _WKBG   = "#FFF0F0"   # 六日背景
    _GRID   = "#E8EBF4"
    _INK    = "#2D3048"
    _RED    = "#CC2200"   # 六日文字

    def __init__(self, parent, **kw):
        kw.setdefault("highlightthickness", 0)
        kw.setdefault("bd", 0)
        kw.setdefault("bg", self._BG)
        kw.setdefault("height", self.TOTAL_H)
        super().__init__(parent, **kw)
        self._dates: list[dt.date] = []

    def set_dates(self, dates: list[dt.date]) -> None:
        self._dates = list(dates)
        self._redraw()

    def sync_x(self, first, *_) -> None:
        """由 Treeview xscrollcommand 呼叫，同步水平捲動位置。"""
        self.xview_moveto(first)

    def _redraw(self) -> None:
        self.delete("all")
        dates   = self._dates
        total_w = self.EMP_W + self.HOURS_W + self.COL_W * len(dates)
        self.configure(scrollregion=(0, 0, total_w, self.TOTAL_H))

        # 員工欄（跨兩排高度）
        self.create_rectangle(0, 0, self.EMP_W, self.TOTAL_H,
                               fill=self._BG, outline=self._GRID)
        self.create_text(self.EMP_W // 2, self.TOTAL_H // 2,
                         text="員工",
                         font=("TkDefaultFont", 12, "bold"),
                         fill=self._INK, anchor="center")

        # 時數欄（跨兩排高度）
        hx = self.EMP_W
        self.create_rectangle(hx, 0, hx + self.HOURS_W, self.TOTAL_H,
                               fill=self._BG, outline=self._GRID)
        self.create_text(hx + self.HOURS_W // 2, self.TOTAL_H // 2,
                         text="時數",
                         font=("TkDefaultFont", 12, "bold"),
                         fill=self._INK, anchor="center")

        for i, d in enumerate(dates):
            x0 = self.EMP_W + self.HOURS_W + i * self.COL_W
            x1 = x0 + self.COL_W
            is_wkend = d.weekday() >= 5   # 5=Sat, 6=Sun
            bg = self._WKBG if is_wkend else self._BG
            fg = self._RED  if is_wkend else self._INK

            # 上列：月/日
            self.create_rectangle(x0, 0, x1, self.DATE_H,
                                   fill=bg, outline=self._GRID)
            self.create_text((x0 + x1) // 2, self.DATE_H // 2,
                             text=f"{d.month}/{d.day}",
                             font=("TkDefaultFont", 11, "bold"),
                             fill=fg, anchor="center")

            # 下列：星期
            self.create_rectangle(x0, self.DATE_H, x1, self.TOTAL_H,
                                   fill=bg, outline=self._GRID)
            self.create_text((x0 + x1) // 2,
                             self.DATE_H + self.WD_H // 2,
                             text=_WEEK[d.weekday()],
                             font=("TkDefaultFont", 11),
                             fill=fg, anchor="center")


class _ScheduleGrid(tk.Frame):
    """Canvas-based schedule grid；kind=3（排休）顯示紅字。"""
    ROW_H   = 26
    EMP_W   = 150
    HOURS_W = 60
    COL_W   = 52
    FONT    = ("TkDefaultFont", 11)
    BG      = ("#FFFFFF", "#F8F9FD")
    GRID_C  = "#E8EBF4"
    INK     = "#2D3048"
    REST_FG = "#CC2200"   # kind=3 排休紅字

    def __init__(self, parent, on_double=None, **kw):
        super().__init__(parent, **kw)
        self._on_double = on_double          # callback(emp_obj, col_idx)
        self._rows: list = []                # [(emp, values_list, kinds_list)]
        self._dates: list = []

        self._cv = tk.Canvas(self, bg="#FFFFFF", highlightthickness=0)
        # 設定捲動單位=列高,讓滾輪/觸控板一格捲一列(預設 0 會導致幾乎不動)
        self._cv.configure(yscrollincrement=self.ROW_H)
        self._cv.grid(row=0, column=0, sticky="nsew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        _bind_mousewheel(self._cv)
        self._cv.bind("<Double-Button-1>", self._handle_double)

    # ── 對外 scroll API（接替 ttk.Treeview） ──────────────────────────────
    def configure(self, **kw):
        ys = kw.pop("yscrollcommand", None)
        xs = kw.pop("xscrollcommand", None)
        if ys: self._cv.configure(yscrollcommand=ys)
        if xs: self._cv.configure(xscrollcommand=xs)
        if kw: super().configure(**kw)

    def yview(self, *a): self._cv.yview(*a)
    def xview(self, *a): self._cv.xview(*a)
    def xview_moveto(self, f): self._cv.xview_moveto(f)

    # ── 填充資料 ──────────────────────────────────────────────────────────
    def populate(self, dates: list, emp_rows: list) -> None:
        """
        dates    : list[dt.date]
        emp_rows : [(emp_obj, values_list, kinds_list)]
          values_list[0] = 員工標籤, [1] = 時數, [2+] = 班別代號
          kinds_list[j]  = 第 j 個日期的 pb29004 (int|None)
        """
        self._cv.delete("all")
        self._rows  = emp_rows
        self._dates = dates
        total_w = self.EMP_W + self.HOURS_W + self.COL_W * len(dates)

        for i, (emp, vals, kinds) in enumerate(emp_rows):
            y0 = i * self.ROW_H
            y1 = y0 + self.ROW_H
            bg = self.BG[i % 2]

            # 員工欄（靠左文字）
            x = 0
            self._cv.create_rectangle(x, y0, x + self.EMP_W, y1,
                                      fill=bg, outline=self.GRID_C)
            self._cv.create_text(x + 8, (y0 + y1) // 2, text=vals[0],
                                 font=self.FONT, fill=self.INK,
                                 anchor="w", width=self.EMP_W - 12)
            x += self.EMP_W

            # 時數欄（置中）
            self._cv.create_rectangle(x, y0, x + self.HOURS_W, y1,
                                      fill=bg, outline=self.GRID_C)
            self._cv.create_text(x + self.HOURS_W // 2, (y0 + y1) // 2,
                                 text=vals[1], font=self.FONT, fill=self.INK,
                                 anchor="center")
            x += self.HOURS_W

            # 日期欄
            for j in range(len(dates)):
                code = vals[2 + j] if 2 + j < len(vals) else ""
                kind = kinds[j] if j < len(kinds) else None
                fg   = self.REST_FG if self._is_rest_cell(kind, code) else self.INK
                self._cv.create_rectangle(x, y0, x + self.COL_W, y1,
                                          fill=bg, outline=self.GRID_C)
                if code:
                    self._cv.create_text(x + self.COL_W // 2, (y0 + y1) // 2,
                                         text=code, font=self.FONT, fill=fg,
                                         anchor="center", width=self.COL_W - 4)
                x += self.COL_W

        total_h = len(emp_rows) * self.ROW_H
        self._cv.configure(scrollregion=(0, 0, total_w, total_h))

    @staticmethod
    def _is_rest_cell(kind, code: str) -> bool:
        """kind=3 → 排休紅字；kind=None 時以 991* 代碼判斷。"""
        if kind == 3:
            return True
        if kind is None and code and code.startswith("991"):
            return True
        return False

    def _handle_double(self, event) -> None:
        if not self._on_double or not self._rows:
            return
        cx = self._cv.canvasx(event.x)
        cy = self._cv.canvasy(event.y)
        if cx < self.EMP_W + self.HOURS_W:
            return   # 點到員工欄或時數欄
        col_idx = int((cx - self.EMP_W - self.HOURS_W) // self.COL_W)
        row_idx = int(cy // self.ROW_H)
        if row_idx < 0 or row_idx >= len(self._rows):
            return
        if col_idx < 0 or col_idx >= len(self._dates):
            return
        self._on_double(self._rows[row_idx][0], col_idx)


class _PunchTable(tk.Frame):
    """Canvas-based table 支援單格著色。"""

    COLS = [
        ("員工代號",  90), ("員工姓名",  90), ("出勤日期", 110),
        ("排班代號",  70), ("表定上班",  80), ("表定下班",  80),
        ("實際上班",  80), ("實際下班",  80), ("備注",     140),
    ]
    ROW_H  = 27
    HDR_H  = 32
    FONT   = ("TkDefaultFont", 12)
    BFONT  = ("TkDefaultFont", 12, "bold")
    HDR_BG = "#EEF2FB"
    BG     = ("#FFFFFF", "#F8F9FD")
    GRID   = "#E8EBF4"

    def __init__(self, parent, **kw):
        super().__init__(parent, **kw)
        self._data: list[tuple] = []
        self._cv = tk.Canvas(self, bg="#FFFFFF", highlightthickness=0)
        # 設定捲動單位=列高,讓滾輪/觸控板一格捲一列(預設 0 會導致幾乎不動)
        self._cv.configure(yscrollincrement=self.ROW_H)
        vsb = ttk.Scrollbar(self, orient="vertical",   command=self._cv.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self._cv.xview)
        self._cv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._cv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)
        _bind_mousewheel(self._cv)
        self._tw = sum(w for _, w in self.COLS)
        self._draw_header()

    def _draw_header(self):
        x = 0
        for name, w in self.COLS:
            self._cv.create_rectangle(x, 0, x+w, self.HDR_H,
                                      fill=self.HDR_BG, outline=self.GRID)
            self._cv.create_text(x + w//2, self.HDR_H//2,
                                 text=name, font=self.BFONT, fill="#2D3048",
                                 anchor="center")
            x += w

    def populate(self, rows: list[tuple]) -> None:
        """rows: list of (values_tuple, fgs_tuple) — fgs 為各格顏色，空字串=黑。"""
        self._cv.delete("row")
        self._data = list(rows)
        y = self.HDR_H
        for i, (vals, fgs) in enumerate(rows):
            bg = self.BG[i % 2]
            x = 0
            for j, (val, (_, w)) in enumerate(zip(vals, self.COLS)):
                fg = fgs[j] if j < len(fgs) and fgs[j] else "#000000"
                self._cv.create_rectangle(x, y, x+w, y+self.ROW_H,
                                          fill=bg, outline=self.GRID, tags="row")
                self._cv.create_text(x+6, y+self.ROW_H//2,
                                     text="" if val is None else str(val),
                                     font=self.FONT, fill=fg,
                                     anchor="w", width=w-10, tags="row")
                x += w
            y += self.ROW_H
        self._cv.configure(scrollregion=(0, 0, self._tw, y))

    def all_values(self) -> list[tuple]:
        return [v for v, _ in self._data]


class _SuggestionTable(tk.Frame):
    """High-density suggestion list with row selection and quick apply."""

    COLS = [
        ("status", "狀態", 58, "center"),
        ("type", "類型", 70, "center"),
        ("date", "日期", 92, "center"),
        ("emp", "員工", 120, "w"),
        ("current", "原班別", 72, "center"),
        ("scheduled", "表定", 92, "center"),
        ("actual", "實際", 92, "center"),
        ("delta", "差異", 92, "center"),
        ("best", "最佳建議", 130, "w"),
        ("hours", "工時", 52, "center"),
        ("action", "套用", 74, "center"),
    ]

    def __init__(self, parent, on_select=None, on_apply=None, **kw):
        super().__init__(parent, **kw)
        self._rows: dict[str, dict] = {}
        self._on_select = on_select
        self._on_apply = on_apply
        cols = [c[0] for c in self.COLS]
        self.tv = ttk.Treeview(
            self, style="App.Treeview", columns=cols, show="headings",
            selectmode="browse", height=18
        )
        for key, label, width, anchor in self.COLS:
            self.tv.heading(key, text=label)
            self.tv.column(key, width=width, anchor=anchor, stretch=False)
        ysb = ttk.Scrollbar(self, orient="vertical", command=self.tv.yview)
        xsb = ttk.Scrollbar(self, orient="horizontal", command=self.tv.xview)
        self.tv.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.tv.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)
        _bind_mousewheel(self.tv)
        self.tv.tag_configure("applied", foreground=_C["green"])
        self.tv.tag_configure("failed", foreground=_C["red"])
        self.tv.tag_configure("empty", foreground=_C["muted"])
        self.tv.bind("<<TreeviewSelect>>", self._handle_select)
        self.tv.bind("<ButtonRelease-1>", self._handle_click)

    def populate(self, rows: list[dict]) -> None:
        self._rows = {r["id"]: r for r in rows}
        self.tv.delete(*self.tv.get_children())
        for row in rows:
            tags = []
            if row.get("status") == "已套用":
                tags.append("applied")
            elif row.get("status") == "失敗":
                tags.append("failed")
            elif not row.get("best_candidate"):
                tags.append("empty")
            self.tv.insert("", "end", iid=row["id"], values=self._values(row),
                           tags=tuple(tags))

    def select(self, row_id: str | None) -> None:
        if row_id and row_id in self._rows:
            self.tv.selection_set(row_id)
            self.tv.focus(row_id)
            self.tv.see(row_id)
            if self._on_select:
                self._on_select(self._rows[row_id])

    def first_id(self) -> str | None:
        children = self.tv.get_children()
        return children[0] if children else None

    def _values(self, row: dict) -> tuple:
        best = row.get("best_candidate")
        best_txt = "無合適建議"
        hours = ""
        action = ""
        if best:
            code, si, so, _d_in, _d_out, h = best
            name = row.get("shift_names", {}).get(code, "")
            best_txt = f"{code} {name or f'{si}-{so}'}".strip()
            hours = f"{h:g}h"
            action = "" if row.get("status") in ("已套用", "套用中") else "套用最佳"
        return (
            row.get("status", "待處理"),
            row.get("type", ""),
            row.get("date", ""),
            f"{row.get('emp_id', '')} {row.get('name', '')}".strip(),
            row.get("current_shift", ""),
            row.get("scheduled", ""),
            row.get("actual", ""),
            row.get("delta", ""),
            best_txt,
            hours,
            action,
        )

    def _handle_select(self, _event=None) -> None:
        sel = self.tv.selection()
        if sel and self._on_select:
            self._on_select(self._rows.get(sel[0]))

    def _handle_click(self, event) -> None:
        iid = self.tv.identify_row(event.y)
        col = self.tv.identify_column(event.x)
        if not iid or col != f"#{len(self.COLS)}":
            return
        row = self._rows.get(iid)
        if (row and row.get("best_candidate")
                and row.get("status") not in ("已套用", "套用中")):
            if self._on_apply:
                self._on_apply(iid)


class EhrsApp(ctk.CTk):
    def __init__(self, auto_login: bool = True) -> None:
        super().__init__()
        self.title(f"文中排班工具  v{__version__}")
        self.geometry("1180x720")
        self.configure(fg_color=_C["app_bg"])

        self.client: EhrsClient | None = None
        self.account: str = ""
        self.cfg = load_config()
        self.schedule: list = []
        self.shift_codes: dict[str, tuple[str, int]] = dict(SHIFT_CODES_REF)
        self._punch_cache: list[tuple] = []
        self._raw_punch_rows: list[dict] = []   # 原始刷卡紀錄（每刷一筆一列）
        self._raw_punch_range: tuple | None = None   # 上次打卡查詢的 (start_iso, end_iso)
        self._selected_emps: set[str] = set()    # 全域篩選：empty = 全員
        self._selected_depts: set[str] = set()   # 全域篩選：empty = 全部門
        self._emp_dept: dict[str, str] = {}      # emp_id -> 部門名稱
        self._emp_dept_code: dict[str, str] = {} # emp_id -> 部門代號

        self._build_ui()
        if auto_login:
            self.after(250, self._do_login)

    # ------------------------------------------------------------------ UI
    def _build_ui(self) -> None:
        # ── 設定 ttk 樣式（排班格 + 打卡紀錄用的 Treeview）──────────
        style = ttk.Style()
        style.configure("App.Treeview.Heading",
                         background=_C["tbl_hdr"], foreground=_C["ink"],
                         font=("TkDefaultFont", 12, "bold"), relief="flat")
        style.configure("App.Treeview",
                         rowheight=28, background=_C["white"],
                         foreground=_C["ink"], fieldbackground=_C["white"],
                         borderwidth=0)
        style.map("App.Treeview",
                  background=[("selected", _C["accent"])],
                  foreground=[("selected", "#FFFFFF")])
        # 去掉 Heading 的 relief border
        style.layout("App.Treeview.Heading", [
            ("Treeheading.cell", {"sticky": "nswe"}),
            ("Treeheading.border", {"sticky": "nswe", "children": [
                ("Treeheading.padding", {"sticky": "nswe", "children": [
                    ("Treeheading.label", {"sticky": "nswe"})
                ]})
            ]}),
        ])

        # ── 頂部品牌列 ─────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color=_C["white"], corner_radius=0,
                            height=54, border_width=0)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)
        # 底線
        tk.Frame(hdr, bg=_C["line"], height=1).pack(side="bottom", fill="x")

        # Logo 方塊
        logo = ctk.CTkFrame(hdr, fg_color=_C["accent"], width=34, height=34,
                              corner_radius=9)
        logo.pack(side="left", padx=(16, 10), pady=10)
        logo.pack_propagate(False)
        ctk.CTkLabel(logo, text="文", font=("", 18, "bold"),
                      text_color="#FFFFFF").pack(expand=True)

        ctk.CTkLabel(hdr, text="文中排班",
                      font=("", 16, "bold"), text_color=_C["ink"]).pack(side="left")
        ctk.CTkLabel(hdr, text=f"v{__version__}",
                      font=("", 11), text_color=_C["muted"]).pack(
            side="left", padx=(5, 0), pady=(7, 0))

        # ── 分頁 ───────────────────────────────────────────────────
        self.tabview = ctk.CTkTabview(
            self,
            fg_color=_C["white"],
            segmented_button_fg_color=_C["tab_bg"],
            segmented_button_selected_color=_C["white"],
            segmented_button_selected_hover_color=_C["white"],
            segmented_button_unselected_color=_C["tab_bg"],
            segmented_button_unselected_hover_color="#DDDFE8",
            text_color=_C["accent"],
            text_color_disabled=_C["muted"],
            border_color=_C["line"],
            border_width=1,
        )
        self.tabview.pack(fill="both", expand=True, padx=10, pady=(8, 4))
        self._build_schedule_tab(self.tabview.add("排班"))
        self._build_punch_tab(self.tabview.add("打卡"))
        self._build_raw_punch_tab(self.tabview.add("打卡紀錄"))
        self._build_suggest_tab(self.tabview.add("建議"))
        self._build_settings_tab(self.tabview.add("設定"))

        # ── 底部 status bar（含匯入進度條）─────────────────────────
        btm = ctk.CTkFrame(self, fg_color=_C["white"], height=36,
                            corner_radius=0, border_width=0)
        btm.pack(side="bottom", fill="x")
        btm.pack_propagate(False)
        # 頂線
        tk.Frame(btm, bg=_C["line"], height=1).pack(side="top", fill="x")

        # 狀態指示燈
        self._status_dot = ctk.CTkLabel(btm, text="●", font=("", 10),
                                          text_color=_C["muted"], width=16)
        self._status_dot.pack(side="left", padx=(14, 0))

        self.status = ctk.CTkLabel(btm, text="就緒", anchor="w",
                                     text_color=_C["muted"], font=("", 12))
        self.status.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # 進度條區塊（平時隱藏，匯入時顯示）
        self._prog_frm = ctk.CTkFrame(btm, fg_color="transparent")
        self._prog_bar = ctk.CTkProgressBar(self._prog_frm, width=180, height=12,
                                              progress_color=_C["accent"],
                                              fg_color=_C["line"])
        self._prog_bar.set(0)
        self._prog_bar.pack(side="left", padx=(0, 6))
        self._prog_lbl = ctk.CTkLabel(self._prog_frm, text="", width=70,
                                        anchor="e", font=("", 12),
                                        text_color=_C["muted"])
        self._prog_lbl.pack(side="left")

    def _build_schedule_tab(self, tab) -> None:
        bar = ctk.CTkFrame(tab, fg_color=_C["white"], corner_radius=8,
                            border_width=1, border_color=_C["line"])
        bar.pack(fill="x", padx=6, pady=6)

        now = dt.date.today()
        self.year_var  = ctk.StringVar(value=str(now.year))
        self.month_var = ctk.StringVar(value=str(now.month))
        self._period_start: dt.date | None = None
        self._period_end:   dt.date | None = None
        self._grid_dates:   list[dt.date]  = []
        # 排班頁部門對照；篩選狀態使用全域 _selected_depts/_selected_emps
        self._sched_emp_dept: dict[str, str] = {}

        # ── 民國年 ──────────────────────────────────────────
        ctk.CTkLabel(bar, text="民國", text_color=_C["ink"],
                      font=("", 13)).pack(side="left", padx=(8, 2), pady=5)
        self.roc_year_var = ctk.StringVar(value=str(now.year - 1911))
        roc_ent = ctk.CTkEntry(bar, textvariable=self.roc_year_var, width=48,
                                border_color=_C["line"], fg_color=_C["white"],
                                text_color=_C["ink"])
        roc_ent.pack(side="left", pady=5)
        ctk.CTkLabel(bar, text="年", text_color=_C["ink"],
                      font=("", 13)).pack(side="left", padx=(2, 6))

        # ── 排班期間下拉 ─────────────────────────────────────
        self.period_var  = ctk.StringVar()
        self.period_menu = ctk.CTkOptionMenu(
            bar, variable=self.period_var, width=156,
            fg_color=_C["white"], button_color=_C["accent"],
            button_hover_color=_C["acc_h"],
            dropdown_fg_color=_C["white"],
            dropdown_hover_color=_C["tab_bg"],
            text_color=_C["ink"], dropdown_text_color=_C["ink"],
            command=self._on_period_select,
        )
        self.period_menu.pack(side="left", pady=5)
        roc_ent.bind("<Return>",   lambda e: self._refresh_period_menu())
        roc_ent.bind("<FocusOut>", lambda e: self._refresh_period_menu())

        # ── 操作按鈕 ─────────────────────────────────────────
        ctk.CTkButton(
            bar, text="載入排班", width=_W_PRIMARY,
            fg_color=_C["accent"], hover_color=_C["acc_h"],
            corner_radius=8, font=("", 13, "bold"),
            command=self._load_schedule,
        ).pack(side="left", padx=(8, 6), pady=5)
        ctk.CTkLabel(bar, text="雙擊改班/刪班",
                      text_color=_C["muted"], font=("", 12)).pack(side="left", padx=2)
        ctk.CTkButton(bar, text="匯出 Excel", width=_W_BTN,
                       fg_color=_C["muted"], hover_color="#6B7280",
                       corner_radius=8, command=self._export_schedule
                       ).pack(side="right", padx=(3, 4), pady=5)
        ctk.CTkButton(bar, text="匯入 Excel", width=_W_BTN,
                       fg_color=_C["muted"], hover_color="#6B7280",
                       corner_radius=8, command=self._import_schedule
                       ).pack(side="right", padx=(3, 0), pady=5)
        ctk.CTkButton(bar, text="特休餘額", width=_W_BTN,
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=8, command=self._show_annual_leave
                       ).pack(side="right", padx=(3, 0), pady=5)
        ctk.CTkButton(bar, text="到職日", width=_W_BTN,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"],
                       corner_radius=8, command=self._show_staff_info
                       ).pack(side="right", padx=(3, 0), pady=5)
        self.sched_emp_btn = ctk.CTkButton(
            bar, text="員工：全員 ▼", width=_W_FILTER,
            fg_color=_C["tab_bg"], text_color=_C["ink"], hover_color=_C["line"],
            corner_radius=8, command=self._open_sched_emp_picker)
        self.sched_emp_btn.pack(side="right", padx=(3, 0), pady=5)
        self.sched_dept_btn = ctk.CTkButton(
            bar, text="部門：全部 ▼", width=_W_FILTER,
            fg_color=_C["tab_bg"], text_color=_C["ink"], hover_color=_C["line"],
            corner_radius=8, command=self._open_sched_dept_picker)
        self.sched_dept_btn.pack(side="right", padx=(3, 0), pady=5)

        # ── 排班格 ───────────────────────────────────────────
        wrap = tk.Frame(tab, bg=_C["app_bg"])
        wrap.pack(fill="both", expand=True, padx=6, pady=(0, 6))

        # 自訂雙排表頭（月/日 + 星期，六日紅字）
        self._sched_header = _ScheduleHeader(wrap)
        self._sched_header.grid(row=0, column=0, sticky="ew")

        self.grid_tv = _ScheduleGrid(wrap, on_double=self._on_grid_double_canvas,
                                      bg=_C["app_bg"])
        ysb = ttk.Scrollbar(wrap, orient="vertical",   command=self.grid_tv.yview)
        xsb = ttk.Scrollbar(wrap, orient="horizontal", command=self.grid_tv.xview)

        def _on_xscroll(first, last):
            xsb.set(first, last)
            self._sched_header.sync_x(first)

        self.grid_tv.configure(yscrollcommand=ysb.set, xscrollcommand=_on_xscroll)
        self.grid_tv.grid(row=1, column=0, sticky="nsew")
        ysb.grid(row=1, column=1, sticky="ns")
        xsb.grid(row=2, column=0, sticky="ew")
        wrap.rowconfigure(1, weight=1)
        wrap.columnconfigure(0, weight=1)

        # 初始化期間選單
        self._refresh_period_menu()

    # ── 排班期間選單輔助 ───────────────────────────────────────────────────────
    def _refresh_period_menu(self) -> None:
        """依目前民國年重建排班期間選單，並自動定位到當前期間。"""
        try:
            roc_year = int(self.roc_year_var.get().strip())
        except ValueError:
            return
        self.year_var.set(str(roc_year + 1911))
        periods = _get_schedule_periods(roc_year)
        labels  = self._period_labels(periods)
        self.period_menu.configure(values=labels)
        # 自動選中包含今天的期間（否則選最後一期）
        today   = dt.date.today()
        sel_idx = len(periods) - 1
        for i, (s, e) in enumerate(periods):
            if s <= today <= e:
                sel_idx = i
                break
        self.period_var.set(labels[sel_idx])
        self._on_period_select(labels[sel_idx])

    def _on_period_select(self, choice: str) -> None:
        """期間下拉變更：更新內部 _period_start / _period_end。"""
        try:
            roc_year = int(self.roc_year_var.get().strip())
        except ValueError:
            return
        periods = _get_schedule_periods(roc_year)
        labels  = self._period_labels(periods)
        try:
            idx = labels.index(choice)
        except ValueError:
            return
        s, e = periods[idx]
        self._period_start = s
        self._period_end   = e
        self.year_var.set(str(s.year))
        self.month_var.set(str(s.month))

    @staticmethod
    def _period_labels(periods: list) -> list[str]:
        return [
            f"第 {i+1:>2} 期　{s.month}/{s.day} ～ {e.month}/{e.day}"
            for i, (s, e) in enumerate(periods)
        ]

    def _build_punch_tab(self, tab) -> None:
        _btn_outline = dict(fg_color=_C["white"], text_color=_C["ink"],
                             hover_color=_C["tab_bg"], border_width=1,
                             border_color=_C["line"], corner_radius=8)
        bar = ctk.CTkFrame(tab, fg_color=_C["white"], corner_radius=8,
                            border_width=1, border_color=_C["line"])
        bar.pack(fill="x", padx=6, pady=6)
        today = dt.date.today()
        self.p_start = ctk.CTkEntry(bar, width=100, border_color=_C["line"],
                                      fg_color=_C["white"], text_color=_C["ink"])
        self.p_start.insert(0, today.replace(day=1).isoformat())
        self.p_end = ctk.CTkEntry(bar, width=100, border_color=_C["line"],
                                    fg_color=_C["white"], text_color=_C["ink"])
        self.p_end.insert(0, today.isoformat())
        self.p_dept_btn = ctk.CTkButton(
            bar, text="部門：全部 ▼", width=_W_FILTER, **_btn_outline,
            command=self._open_dept_picker
        )
        self.p_emp_btn = ctk.CTkButton(
            bar, text="員工：全員 ▼", width=_W_FILTER, **_btn_outline,
            command=self._open_emp_picker
        )
        self.p_show_err   = ctk.CTkCheckBox(bar, text="異常",
                                             text_color=_C["ink"], fg_color=_C["red"],
                                             command=self._apply_punch_filter)
        self.p_show_ot    = ctk.CTkCheckBox(bar, text="加班",
                                             text_color=_C["ink"], fg_color="#C07000",
                                             command=self._apply_punch_filter)
        self.p_show_leave = ctk.CTkCheckBox(bar, text="請假",
                                             text_color=_C["ink"], fg_color=_C["green"],
                                             command=self._apply_punch_filter)
        self.p_show_err.select()
        self.p_show_ot.select()
        self.p_show_leave.select()
        ctk.CTkLabel(bar, text="起", text_color=_C["ink"],
                      font=("", 13)).pack(side="left", padx=(8, 2))
        self.p_start.pack(side="left", pady=5)
        ctk.CTkLabel(bar, text="迄", text_color=_C["ink"],
                      font=("", 13)).pack(side="left", padx=(8, 2))
        self.p_end.pack(side="left", pady=5)
        self.p_dept_btn.pack(side="left", padx=(8, 0), pady=5)
        self.p_emp_btn.pack(side="left", padx=(4, 0), pady=5)
        self.p_show_err.pack(side="left", padx=(8, 2))
        self.p_show_ot.pack(side="left", padx=2)
        self.p_show_leave.pack(side="left", padx=(2, 4))
        ctk.CTkButton(bar, text="查詢打卡", width=_W_PRIMARY,
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=8, font=("", 13, "bold"),
                       command=self._query_punch).pack(side="left", padx=(8, 4), pady=5)
        ctk.CTkButton(bar, text="匯出 Excel", width=_W_BTN,
                       fg_color=_C["muted"], hover_color="#6B7280",
                       corner_radius=8, command=self._export_punch
                       ).pack(side="right", padx=(3, 4), pady=5)
        ctk.CTkButton(bar, text="出勤統計", width=_W_BTN,
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=8, command=self._show_attendance_stats
                       ).pack(side="right", padx=(0, 3), pady=5)

        wrap = tk.Frame(tab, bg=_C["app_bg"])
        wrap.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        self.punch_tbl = _PunchTable(wrap)
        self.punch_tbl.grid(row=0, column=0, sticky="nsew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

    def _build_raw_punch_tab(self, tab) -> None:
        _btn_outline = dict(fg_color=_C["white"], text_color=_C["ink"],
                             hover_color=_C["tab_bg"], border_width=1,
                             border_color=_C["line"], corner_radius=8)
        bar = ctk.CTkFrame(tab, fg_color=_C["white"], corner_radius=8,
                            border_width=1, border_color=_C["line"])
        bar.pack(fill="x", padx=6, pady=6)

        self.raw_dept_btn = ctk.CTkButton(
            bar, text="部門：全部 ▼", width=_W_FILTER, **_btn_outline,
            command=self._open_raw_dept_picker
        )
        self.raw_emp_btn = ctk.CTkButton(
            bar, text="員工：全員 ▼", width=_W_FILTER, **_btn_outline,
            command=self._open_raw_emp_picker
        )
        self.raw_count_lbl = ctk.CTkLabel(bar, text="共 0 筆",
                                            text_color=_C["muted"], font=("", 12))

        self.raw_dept_btn.pack(side="left", padx=(8, 0), pady=5)
        self.raw_emp_btn.pack(side="left", padx=(4, 0), pady=5)
        self.raw_count_lbl.pack(side="left", padx=8)
        ctk.CTkButton(bar, text="匯出 Excel", width=_W_BTN,
                       fg_color=_C["muted"], hover_color="#6B7280",
                       corner_radius=8, command=self._export_raw_punch
                       ).pack(side="right", padx=(3, 4), pady=5)

        wrap = tk.Frame(tab, bg=_C["app_bg"])
        wrap.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        cols = ("dept", "emp_id", "emp_name", "date", "time", "jiezhuan", "source")
        self.raw_tv = ttk.Treeview(wrap, style="App.Treeview",
                                    columns=cols, show="headings", height=20)
        for col, lbl, w in [
            ("dept",      "部門",   90),
            ("emp_id",    "員工代號", 80),
            ("emp_name",  "員工姓名", 80),
            ("date",      "出勤日期", 110),
            ("time",      "刷卡時間", 90),
            ("jiezhuan",  "結轉",    60),
            ("source",    "來源",    70),
        ]:
            self.raw_tv.heading(col, text=lbl)
            self.raw_tv.column(col, width=w, anchor="center", stretch=False)
        ysb = ttk.Scrollbar(wrap, orient="vertical",   command=self.raw_tv.yview)
        xsb = ttk.Scrollbar(wrap, orient="horizontal", command=self.raw_tv.xview)
        self.raw_tv.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.raw_tv.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)
        _bind_mousewheel(self.raw_tv)

    def _populate_raw_punch(self) -> None:
        rows = self._raw_punch_rows
        # 套用篩選
        def _match(r: dict) -> bool:
            eid  = _first_field(r, _EMP_ID_KEYS)
            dept = self._emp_filter_dept(eid, row=r)
            if not self._dept_passes_filter(dept):
                return False
            if self._selected_emps and eid not in self._selected_emps:
                return False
            return True

        filtered = [r for r in rows if _match(r)]
        self.raw_tv.delete(*self.raw_tv.get_children())
        for r in filtered:
            eid    = _first_field(r, _EMP_ID_KEYS)
            dept   = self._emp_filter_dept(eid, row=r)
            ename  = _first_field(r, _EMP_NAME_KEYS)
            date   = str(r.get("出勤日期", ""))[:10]
            time_  = str(r.get("刷卡時間", ""))[11:16]
            jz     = r.get("結轉註記", "")
            src    = r.get("來源", "")
            self.raw_tv.insert("", "end", values=(dept, eid, ename, date, time_, jz, src))
        self.raw_count_lbl.configure(text=f"共 {len(filtered)} 筆")

    # ── 打卡紀錄篩選器 ────────────────────────────────────────
    def _open_raw_dept_picker(self) -> None:
        depts = sorted({
            self._emp_filter_dept(_first_field(r, _EMP_ID_KEYS), row=r)
            for r in self._raw_punch_rows
        } - {""})
        if not depts:
            messagebox.showinfo("提示", "請先在打卡分頁查詢資料。")
            return
        self._generic_picker(
            title="選擇部門", items=depts,
            selected=self._selected_depts,
            on_confirm=lambda s: self._set_global_filter(depts=s)
        )

    def _open_raw_emp_picker(self) -> None:
        emps: dict[str, str] = {}
        for r in self._raw_punch_rows:
            eid = _first_field(r, _EMP_ID_KEYS)
            name = _first_field(r, _EMP_NAME_KEYS)
            # 只顯示符合部門篩選的員工
            dept = self._emp_filter_dept(eid, row=r)
            if not self._dept_passes_filter(dept):
                continue
            if eid:
                emps[eid] = name
        if not emps:
            messagebox.showinfo("提示", "請先在打卡分頁查詢資料。")
            return
        self._generic_picker(
            title="選擇員工",
            items=[f"{eid}  {name}" for eid, name in sorted(emps.items())],
            selected={f"{eid}  {emps[eid]}" for eid in self._selected_emps if eid in emps},
            on_confirm=lambda s: self._set_global_filter(
                emps={item.split()[0] for item in s}
            ),
            searchable=True,
        )

    def _generic_picker(
        self, title: str, items: list[str], selected: set[str],
        on_confirm, searchable: bool = False
    ) -> None:
        """通用多選清單彈窗。"""
        top = ctk.CTkToplevel(self)
        top.title(title)
        top.geometry("280x460")
        top.configure(fg_color=_C["app_bg"])
        top.resizable(False, True)
        top.transient(self)
        top.after(60, top.grab_set)

        search_var = ctk.StringVar()
        if searchable:
            ctk.CTkEntry(top, placeholder_text="搜尋…",
                          textvariable=search_var,
                          border_color=_C["line"], fg_color=_C["white"],
                          text_color=_C["ink"]).pack(
                fill="x", padx=12, pady=(12, 4))

        scroll = ctk.CTkScrollableFrame(top, fg_color=_C["white"],
                                          corner_radius=8,
                                          border_width=1, border_color=_C["line"])
        _bind_scrollable_frame_mousewheel(scroll)
        scroll.pack(fill="both", expand=True, padx=12,
                     pady=(8 if not searchable else 4, 0))

        chk_vars: dict[str, tk.BooleanVar] = {}
        chk_widgets: dict[str, ctk.CTkCheckBox] = {}
        for item in items:
            init = (not selected) or (item in selected)
            var = tk.BooleanVar(value=init)
            chk = ctk.CTkCheckBox(scroll, text=item, variable=var,
                                   text_color=_C["ink"])
            chk.pack(anchor="w", padx=4, pady=3)
            chk_vars[item] = var
            chk_widgets[item] = chk

        if searchable:
            def _filter(*_):
                q = search_var.get().lower()
                for it, w in chk_widgets.items():
                    if not q or q in it.lower():
                        w.pack(anchor="w", padx=4, pady=3)
                    else:
                        w.pack_forget()
            search_var.trace_add("write", _filter)

        btn_bar = ctk.CTkFrame(top, fg_color="transparent")
        btn_bar.pack(fill="x", padx=12, pady=(6, 0))
        ctk.CTkButton(btn_bar, text="全選", width=80,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"], corner_radius=8,
                       command=lambda: [v.set(True)  for v in chk_vars.values()]
                       ).pack(side="left", padx=(0, 4))
        ctk.CTkButton(btn_bar, text="清除", width=80,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"], corner_radius=8,
                       command=lambda: [v.set(False) for v in chk_vars.values()]
                       ).pack(side="left")

        def apply():
            picked = {it for it, v in chk_vars.items() if v.get()}
            final  = set() if len(picked) == len(items) else picked
            on_confirm(final)
            top.destroy()

        ctk.CTkButton(top, text="確定",
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=9, font=("", 13, "bold"),
                       command=apply).pack(fill="x", padx=12, pady=10)

    def _export_raw_punch(self) -> None:
        rows = list(self.raw_tv.get_children())
        if not rows:
            messagebox.showwarning("無資料", "沒有打卡紀錄可匯出。")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 活頁簿", "*.xlsx")],
            initialfile="打卡紀錄.xlsx",
        )
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "打卡紀錄"
        headers = ["部門", "員工代號", "員工姓名", "出勤日期", "刷卡時間", "結轉", "來源"]
        ws.append(headers)
        hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = hdr_fill
        for iid in rows:
            ws.append(list(self.raw_tv.item(iid)["values"]))
        for i, w in enumerate([12, 10, 10, 14, 10, 8, 8], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        wb.save(path)
        self._set_status(f"已匯出打卡紀錄 → {path}")

    def _build_settings_tab(self, tab) -> None:
        # 垂直置中：上下各放一個彈性區塊
        ctk.CTkFrame(tab, fg_color="transparent").pack(fill="both", expand=True)

        card = ctk.CTkFrame(tab, fg_color=_C["white"], corner_radius=14,
                             border_width=1, border_color=_C["line"])
        card.pack(padx=60, anchor="center")

        # ── 卡片標題 ──────────────────────────────────────────
        hdr_row = ctk.CTkFrame(card, fg_color="transparent")
        hdr_row.pack(fill="x", padx=28, pady=(24, 0))

        icon_box = ctk.CTkFrame(hdr_row, fg_color=_C["tab_bg"],
                                  width=46, height=46, corner_radius=12)
        icon_box.pack(side="left", padx=(0, 14))
        icon_box.pack_propagate(False)
        ctk.CTkLabel(icon_box, text="⚙", font=("", 20),
                      text_color=_C["accent"]).pack(expand=True)

        title_col = ctk.CTkFrame(hdr_row, fg_color="transparent")
        title_col.pack(side="left")
        ctk.CTkLabel(title_col, text="連線設定",
                      font=("", 17, "bold"), text_color=_C["ink"],
                      anchor="w").pack(anchor="w")
        ctk.CTkLabel(title_col, text="帳密只存本機 ~/.ehrs_tool，不會上傳",
                      font=("", 12), text_color=_C["muted"],
                      anchor="w").pack(anchor="w")

        # 分隔線
        tk.Frame(card, bg=_C["line"], height=1).pack(fill="x", pady=(18, 0))

        # ── 表單 ──────────────────────────────────────────────
        form = ctk.CTkFrame(card, fg_color="transparent")
        form.pack(fill="x", padx=28, pady=(18, 26))

        def _field(label: str, attr: str, default: str, **kw) -> None:
            ctk.CTkLabel(form, text=label,
                          font=("", 13, "bold"), text_color=_C["ink"],
                          anchor="w").pack(anchor="w", pady=(0, 4))
            ent = ctk.CTkEntry(form, width=400, height=42, corner_radius=9,
                                border_color=_C["line"], fg_color=_C["white"],
                                text_color=_C["ink"], font=("", 13), **kw)
            ent.insert(0, default)
            ent.pack(anchor="w", pady=(0, 14))
            setattr(self, attr, ent)

        _field("帳號",   "s_acc",  self.cfg.get("account",  ""))
        _field("密碼",   "s_pwd",  self.cfg.get("password", ""), show="*")
        _field("站台網址", "s_base", self.cfg.get("base_url", DEFAULT_BASE_URL))

        ctk.CTkButton(
            form, text="儲存並登入", width=400, height=44,
            fg_color=_C["accent"], hover_color=_C["acc_h"],
            corner_radius=10, font=("", 14, "bold"),
            command=self._save_and_login,
        ).pack(anchor="w")

        ctk.CTkFrame(tab, fg_color="transparent").pack(fill="both", expand=True)

    # -------------------------------------------------------------- helpers
    def _set_status(self, text: str) -> None:
        self.status.configure(text=text)
        if not hasattr(self, "_status_dot"):
            return
        # 依訊息內容切換狀態燈顏色
        if any(k in text for k in ("錯誤", "失敗")):
            dot_clr = _C["red"]
        elif any(k in text for k in ("完成", "已載入", "已登入", "已匯出",
                                      "已寫入", "已刪除", "已套用", "安裝完成")):
            dot_clr = _C["green"]
        elif any(k in text for k in ("中…", "查詢", "載入", "匯入", "下載")):
            dot_clr = _C["accent"]
        else:
            dot_clr = _C["muted"]
        self._status_dot.configure(text_color=dot_clr)

    def _show_progress(self, done: int, total: int) -> None:
        """顯示匯入進度條（在主執行緒呼叫）。"""
        if total == 0:
            return
        self._prog_bar.set(done / total)
        self._prog_lbl.configure(text=f"{done} / {total}")
        if not self._prog_frm.winfo_ismapped():
            self._prog_frm.pack(side="right")

    def _hide_progress(self) -> None:
        """隱藏進度條。"""
        self._prog_frm.pack_forget()
        self._prog_bar.set(0)

    def _run_async(self, work, on_done, on_error=None) -> None:
        def worker():
            try:
                result = work()
            except Exception as exc:  # noqa: BLE001 - 回報到 UI
                self.after(0, lambda e=exc: (on_error or self._error)(e))
                return
            self.after(0, lambda: on_done(result))

        threading.Thread(target=worker, daemon=True).start()

    def _error(self, exc: Exception) -> None:
        self._set_status(f"錯誤:{exc}")
        messagebox.showerror("錯誤", str(exc))

    def _need_client(self) -> bool:
        if self.client is None:
            messagebox.showwarning("尚未登入", "請先到「設定」分頁登入。")
            self.tabview.set("設定")
            return False
        return True

    # -------------------------------------------------------------- 自動更新
    def _check_update(self) -> None:
        try:
            from updater import start_update_check
            start_update_check(
                status_cb=lambda msg: self.after(0, lambda m=msg: self._set_status(m))
            )
        except Exception:
            pass  # updater 模組不存在或匯入失敗時靜默略過

    # ---------------------------------------------------------------- login
    def _save_and_login(self) -> None:
        self.cfg = {
            "account": self.s_acc.get().strip(),
            "password": self.s_pwd.get(),
            "base_url": self.s_base.get().strip() or DEFAULT_BASE_URL,
        }
        save_config(self.cfg)
        self._do_login()

    def _do_login(self) -> None:
        acc = self.cfg.get("account")
        pwd = self.cfg.get("password")
        base = self.cfg.get("base_url", DEFAULT_BASE_URL)
        if not acc or not pwd:
            self._set_status("尚未設定帳密,請到「設定」分頁")
            self.tabview.set("設定")
            return
        self._set_status("登入中…")

        def work():
            return EhrsClient(base_url=base).login(acc, pwd)

        def done(client):
            self.client = client
            self.account = acc
            self._set_status(f"已登入:{acc}")

        self._run_async(work, done)

    # ------------------------------------------------------------- schedule
    def _load_schedule(self) -> None:
        if not self._need_client():
            return

        if self._period_start and self._period_end:
            # ── 四週期間模式 ─────────────────────────────────────
            start, end = self._period_start, self._period_end
            period_lbl = f"{start.month}/{start.day}～{end.month}/{end.day}"
            dates = [start + dt.timedelta(days=i)
                     for i in range((end - start).days + 1)]

            # 收集期間橫跨的所有年月
            months_needed: set[tuple[int, int]] = set()
            cur = start.replace(day=1)
            while cur <= end:
                months_needed.add((cur.year, cur.month))
                cur = (cur.replace(day=28) + dt.timedelta(days=4)).replace(day=1)

            self._set_status(f"載入 {period_lbl} 排班中…")

            def work_p():
                results, failures = self._fetch_month_schedules(months_needed)
                emp_map: dict[str, object] = {}
                for ym in sorted(results):           # 早的月份先處理(維持原合併順序)
                    for emp in results[ym]:
                        if emp.emp_id not in emp_map:
                            emp_map[emp.emp_id] = emp
                        else:
                            existing = {s.date for s in emp_map[emp.emp_id].shifts}
                            for s in emp.shifts:
                                if s.date not in existing:
                                    emp_map[emp.emp_id].shifts.append(s)
                return list(emp_map.values()), dates, self._fetch_emp_dept_map(start, end), failures

            def done_p(result):
                sched, dates, dept_map, failures = result
                self.schedule = sched
                for eid, dept in dept_map.items():
                    self._remember_emp_dept(eid, dept)
                self._refresh_sched_emp_dept(sched)
                self._build_code_map(sched)
                self._populate_grid_period(dates, sched)
                self._set_status(
                    f"已載入 {period_lbl} 排班（共 {len(sched)} 人）"
                    + self._fmt_month_failures(failures))

            self._run_async(work_p, done_p)
        else:
            # ── 單月備用模式 ─────────────────────────────────────
            try:
                year  = int(self.year_var.get())
                month = int(self.month_var.get())
            except ValueError:
                messagebox.showwarning("輸入錯誤", "年/月請填數字。")
                return
            self._set_status(f"載入 {year}-{month:02d} 排班中…")

            def work():
                first = dt.date(year, month, 1)
                last = dt.date(year, month, _cal.monthrange(year, month)[1])
                return (
                    self.client.get_schedule(year, month),
                    self._fetch_emp_dept_map(first, last),
                )

            def done(result):
                sched, dept_map = result
                self.schedule = sched
                for eid, dept in dept_map.items():
                    self._remember_emp_dept(eid, dept)
                self._refresh_sched_emp_dept(sched)
                self._build_code_map(sched)
                self._populate_grid(year, month, sched)
                self._set_status(f"已載入 {year}-{month:02d} 排班（共 {len(sched)} 人）")

            self._run_async(work, done)

    def _fetch_month_schedules(self, months):
        """並行抓多個 (year, month) 排班(沿用 set_shifts_bulk 的並行做法)。
        回傳 (results: dict[(y,m)->list[EmployeeSchedule]], failures: list[((y,m), err)])。
        單月失敗不中斷其他月,失敗原因收集回傳供 UI 提示(取代原本無聲 except: pass)。"""
        months = sorted(set(months))
        results: dict[tuple, list] = {}
        failures: list[tuple] = []
        lock = threading.Lock()

        def _w(ym):
            y, m = ym
            try:
                sched = self.client.get_schedule(y, m)
                with lock:
                    results[ym] = sched
            except Exception as exc:  # noqa: BLE001
                with lock:
                    failures.append((ym, str(exc)))

        if months:
            with ThreadPoolExecutor(max_workers=min(6, len(months))) as pool:
                list(pool.map(_w, months))
        return results, failures

    @staticmethod
    def _fmt_month_failures(failures) -> str:
        """把載入失敗的年月組成提示字串;無失敗回空字串。"""
        if not failures:
            return ""
        ms = "、".join(f"{y}/{m:02d}" for (y, m), _ in failures)
        return f"（{ms} 載入失敗，結果可能不完整）"

    def _build_code_map(self, sched) -> None:
        for emp in sched:
            for s in emp.shifts:
                if s.code:
                    existing = self.shift_codes.get(s.code, ("", 3, 0.0))
                    self.shift_codes[s.code] = (
                        s.name or existing[0],
                        s.kind if s.kind is not None else 3,
                        existing[2] if len(existing) > 2 else 0.0,
                    )

    def _populate_grid(self, year: int, month: int, sched) -> None:
        """單月排班格（備用，import 結束後呼叫）。"""
        self._grid_dates = []
        ndays = _cal.monthrange(year, month)[1]
        month_dates = [dt.date(year, month, d) for d in range(1, ndays + 1)]
        self._sched_header.set_dates(month_dates)
        emp_rows = []
        for emp in sched:
            if not self._emp_passes_filter(emp.emp_id, emp):
                continue
            by_day      = {int(s.date[8:10]): s.code for s in emp.shifts}
            by_day_kind = {int(s.date[8:10]): s.kind for s in emp.shifts}
            total_h = sum(
                _shift_hours(by_day[d]) for d in range(1, ndays + 1)
                if d in by_day and by_day[d] and not _is_rest(by_day[d])
            )
            h_str  = f"{total_h:g}" if total_h else ""
            values = [f"{emp.emp_id} {emp.name}", h_str] + [
                by_day.get(d, "") for d in range(1, ndays + 1)
            ]
            kinds  = [by_day_kind.get(d) for d in range(1, ndays + 1)]
            emp_rows.append((emp, values, kinds))
        self.grid_tv.populate(month_dates, emp_rows)

    def _populate_grid_period(self, dates: list[dt.date], sched) -> None:
        """四週期間排班格：以連續日期清單建立 28 欄。"""
        self._grid_dates = list(dates)
        self._sched_header.set_dates(dates)
        emp_rows = []
        for emp in sched:
            if not self._emp_passes_filter(emp.emp_id, emp):
                continue
            by_date      = {s.date: s.code for s in emp.shifts}
            by_date_kind = {s.date: s.kind for s in emp.shifts}
            total_h = sum(
                _shift_hours(by_date[d.isoformat()]) for d in dates
                if d.isoformat() in by_date
                and by_date[d.isoformat()]
                and not _is_rest(by_date[d.isoformat()])
            )
            h_str  = f"{total_h:g}" if total_h else ""
            values = [f"{emp.emp_id} {emp.name}", h_str] + [
                by_date.get(d.isoformat(), "") for d in dates
            ]
            kinds  = [by_date_kind.get(d.isoformat()) for d in dates]
            emp_rows.append((emp, values, kinds))
        self.grid_tv.populate(dates, emp_rows)

    def _on_grid_double_canvas(self, emp, col_idx: int) -> None:
        if self._grid_dates:
            if col_idx >= len(self._grid_dates):
                return
            date_str = self._grid_dates[col_idx].isoformat()
        else:
            day   = col_idx + 1
            year  = int(self.year_var.get())
            month = int(self.month_var.get())
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
        self._open_editor(emp, date_str)

    def _open_editor(self, emp, date_str: str) -> None:
        cur = emp.shift_on(date_str)

        top = ctk.CTkToplevel(self)
        top.title("編輯排班")
        top.geometry("400x530")
        top.configure(fg_color=_C["app_bg"])
        top.transient(self)
        top.after(60, top.grab_set)

        # 員工資訊列
        hdr_row = ctk.CTkFrame(top, fg_color=_C["white"], corner_radius=10,
                                 border_width=1, border_color=_C["line"])
        hdr_row.pack(fill="x", padx=16, pady=(16, 0))
        ctk.CTkLabel(hdr_row, text=f"{emp.emp_id} {emp.name}",
                      font=("", 15, "bold"), text_color=_C["ink"]).pack(
            side="left", padx=14, pady=10)
        ctk.CTkLabel(hdr_row, text=date_str,
                      font=("", 13), text_color=_C["muted"]).pack(
            side="left", padx=4)
        cur_txt = f"目前：{cur.code} {cur.name}" if cur else "目前：（空白）"
        ctk.CTkLabel(hdr_row, text=cur_txt,
                      font=("", 12), text_color=_C["muted"]).pack(
            side="right", padx=14)

        # 預先計算所有班別的上班/下班/時數
        opt_rows: list[tuple[str, str, str, str]] = []  # (code, start, end, hours)
        for c in sorted(self.shift_codes):
            si, so = _shift_times(c)
            h = _shift_hours(c) or (_calc_hours(si, so) if si and so else 0.0)
            h_str = f"{h:g}" if h else ""
            opt_rows.append((c, si, so, h_str))

        entry = ctk.CTkEntry(top, width=368, placeholder_text="輸入班別代號…",
                               border_color=_C["line"], fg_color=_C["white"],
                               text_color=_C["ink"])
        entry.pack(pady=(10, 4), padx=16)
        if cur:
            entry.insert(0, cur.code)

        # ── 排班日種類 ────────────────────────────────────────
        # 預設：991 系列(含 991-1)→ 3(排休)，其他 → 1(工作日)；
        # 若目前已有班別則沿用其 kind。
        _dk = (cur.kind if cur is not None and cur.kind is not None
               else (3 if _is_rest(cur.code if cur else "") else 1))
        kind_var = tk.IntVar(value=_dk)
        kind_bar = ctk.CTkFrame(top, fg_color=_C["tab_bg"], corner_radius=8)
        kind_bar.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(kind_bar, text="排班日種類：",
                      text_color=_C["ink"], font=("", 12)
                      ).pack(side="left", padx=(10, 6), pady=6)
        for _lbl, _val in [("工作日 (1)", 1), ("排休 (3)", 3)]:
            ctk.CTkRadioButton(kind_bar, text=_lbl, variable=kind_var, value=_val,
                                text_color=_C["ink"], font=("", 12)
                                ).pack(side="left", padx=8, pady=6)

        tv_wrap = tk.Frame(top, bg=_C["app_bg"])
        tv_wrap.pack(fill="x", padx=16, pady=(0, 4))
        tv = ttk.Treeview(tv_wrap, style="App.Treeview",
                           columns=("code", "start", "end", "hours"),
                           show="headings", height=10, selectmode="browse")
        tv.heading("code",  text="班別代號"); tv.column("code",  width=90, anchor="center")
        tv.heading("start", text="上班時間"); tv.column("start", width=85, anchor="center")
        tv.heading("end",   text="下班時間"); tv.column("end",   width=85, anchor="center")
        tv.heading("hours", text="時數");     tv.column("hours", width=60, anchor="center")
        sb = ttk.Scrollbar(tv_wrap, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb.set)
        tv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        _bind_mousewheel(tv)

        def _fill(typed: str = "") -> None:
            tv.delete(*tv.get_children())
            t = typed.lower()
            if not t or any(k in t for k in ("空白", "刪除")):
                tv.insert("", "end", iid="__del__", values=("(空白/刪除)", "", "", ""))
            for code, si, so, h in opt_rows:
                name = self.shift_codes.get(code, ("", 3))[0]
                if not t or t in code.lower() or t in name.lower() or t in si or t in so:
                    tv.insert("", "end", values=(code, si, so, h))

        _fill()

        def _pick() -> None:
            sel = tv.selection()
            if sel:
                code = str(tv.item(sel[0])["values"][0])
                entry.delete(0, "end")
                if code != "(空白/刪除)":
                    entry.insert(0, code)

        def _on_entry_key(event) -> None:
            if event.keysym == "Down":
                tv.focus_set()
                children = tv.get_children()
                if children:
                    tv.selection_set(children[0])
                    tv.focus(children[0])
                return
            if event.keysym == "Return":
                save(); return
            _fill(entry.get())

        def _on_tv_key(event) -> None:
            if event.keysym == "Return":
                _pick(); save()
            elif event.keysym == "Up":
                sel = tv.selection()
                children = tv.get_children()
                if sel and children and sel[0] == children[0]:
                    entry.focus_set()

        entry.bind("<KeyRelease>", _on_entry_key)
        tv.bind("<<TreeviewSelect>>", lambda e: _pick())
        tv.bind("<KeyRelease>", _on_tv_key)
        entry.focus_set()

        def save() -> None:
            raw = entry.get().strip()
            top.destroy()
            if not raw or raw.startswith("(空白"):
                self._apply_delete(emp, date_str, cur)
            else:
                self._apply_set(emp, date_str, raw.split()[0], kind_var.get())

        btn_row = ctk.CTkFrame(top, fg_color="transparent")
        btn_row.pack(fill="x", padx=16, pady=(0, 16))
        ctk.CTkButton(btn_row, text="取消", width=170,
                       fg_color=_C["line"], hover_color="#D0D4E4",
                       text_color=_C["ink"], corner_radius=9,
                       command=top.destroy).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_row, text="儲存", width=170,
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=9, font=("", 13, "bold"),
                       command=save).pack(side="left")

    def _apply_set(self, emp, date_str: str, code: str, kind: int = 1) -> None:
        year, month, _ = (int(x) for x in date_str.split("-"))
        self._set_status("寫入中…")
        # kind 由 UI 選擇器傳入（工作日=1 / 排休=3）

        def work():
            return self.client.set_shift(
                year, month, emp.emp_id, date_str, code, kind=kind, dry_run=False
            )

        def done(res):
            self._set_status(f"已寫入 {emp.emp_id} {date_str} → {code}")
            self._load_schedule()

        self._run_async(work, done)

    def _apply_delete(self, emp, date_str: str, cur) -> None:
        if cur is None:
            return
        year, month, _ = (int(x) for x in date_str.split("-"))
        self._set_status("刪除中…")

        def work():
            return self.client.delete_shift(
                year, month, emp.emp_id, date_str, dry_run=False
            )

        def done(res):
            self._set_status(f"已刪除 {emp.emp_id} {date_str}")
            self._load_schedule()

        self._run_async(work, done)

    # --------------------------------------------------------- export / import
    def _export_schedule(self) -> None:
        if not self.schedule:
            messagebox.showwarning("尚未載入", "請先載入排班。")
            return
        # 決定要匯出的日期清單
        if self._grid_dates:
            dates_obj = self._grid_dates
            s, e = dates_obj[0], dates_obj[-1]
            default_name = f"排班_{s.year}-{s.month:02d}{s.day:02d}_{e.month:02d}{e.day:02d}.xlsx"
            sheet_title  = _excel_sheet_title(f"{s.month}-{s.day}～{e.month}-{e.day}")
        else:
            year, month = int(self.year_var.get()), int(self.month_var.get())
            ndays       = _cal.monthrange(year, month)[1]
            dates_obj   = [dt.date(year, month, d) for d in range(1, ndays + 1)]
            default_name = f"排班_{year}-{month:02d}.xlsx"
            sheet_title  = _excel_sheet_title(f"{year}-{month:02d}")

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 活頁簿", "*.xlsx")],
            initialfile=default_name,
        )
        if not path:
            return

        try:
            dates_iso = [d.isoformat() for d in dates_obj]
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_title
            ws.append(["員工代號", "員工姓名", "總時數"] + dates_iso)
            hdr_fill   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            wkend_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
            for cell in ws[1]:
                cell.font      = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.fill      = hdr_fill
            ws.row_dimensions[1].height = 36
            for col_i, d in enumerate(dates_obj, start=4):   # 日期欄從第 4 欄開始
                if d.weekday() >= 5:
                    ws.cell(1, col_i).fill = wkend_fill
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 8    # 總時數
            for i in range(4, len(dates_obj) + 4):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 7
            for emp in self.schedule:
                if not self._emp_passes_filter(emp.emp_id, emp):
                    continue
                by_day = {s.date: s.code for s in emp.shifts}
                total_h = sum(
                    _shift_hours(by_day[d]) for d in dates_iso
                    if d in by_day and by_day[d] and not _is_rest(by_day[d])
                )
                h_str = f"{total_h:g}" if total_h else ""
                ws.append([emp.emp_id, emp.name, h_str] + [by_day.get(d, "") for d in dates_iso])
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                row[0].alignment = Alignment(horizontal="left")
                row[1].alignment = Alignment(horizontal="left")
            wb.save(path)
        except Exception as exc:
            self._set_status("匯出排班失敗")
            messagebox.showerror("匯出失敗", str(exc))
            return
        self._set_status(f"已匯出排班 → {path}")

    def _export_punch(self) -> None:
        data = self.punch_tbl.all_values()
        if not data:
            messagebox.showwarning("無資料", "請先查詢打卡。")
            return
        start = self.p_start.get().strip()
        end = self.p_end.get().strip()
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 活頁簿", "*.xlsx")],
            initialfile=f"打卡_{start}_{end}.xlsx",
        )
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "打卡明細"
        headers = ["員工代號", "員工姓名", "出勤日期", "排班代號", "表定上班", "表定下班", "實際上班", "實際下班", "備注"]
        ws.append(headers)
        hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = hdr_fill
        for row in data:
            ws.append(list(row))
        for i, w in enumerate([10, 10, 14, 10, 10, 10, 10, 10, 18], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        wb.save(path)
        self._set_status(f"已匯出 {len(data)} 筆打卡 → {path}")

    # ── 出勤統計 ───────────────────────────────────────────────────────────────
    def _compute_attendance_stats(self) -> list[tuple]:
        """從 _punch_cache 算每人出勤摘要。"""
        stats: dict[str, dict] = {}
        for vals, _fgs, meta in self._punch_cache:
            emp_id, name, date, code, sched_in, sched_out, clock_in, clock_out, remark = vals
            if emp_id not in stats:
                stats[emp_id] = dict(name=name, sched=0, work=0,
                                     total_h=0.0, ot_h=0.0,
                                     leave=0, absent=0, late=0, early=0)
            s = stats[emp_id]
            remark = remark or ""
            is_err   = any(k in remark for k in _REMARK_ERR_KEYWORDS)
            is_ot    = _REMARK_OT_KEYWORD in remark
            is_leave = bool(remark) and not is_err and not is_ot
            if code and not _is_rest(code):
                s['sched'] += 1
            if clock_in:
                s['work'] += 1
            if clock_in and clock_out:
                ci = _punch_mins(clock_in)
                co = _punch_mins(clock_out)
                if co < ci: co += 1440
                h = (co - ci) / 60
                if h > 9:   h -= 1
                elif h > 4: h -= 0.5
                s['total_h'] += max(0, h)
            # 加班時數直接讀 meta(由 _overtime_hours 單一來源算出),不再從備注字串反解
            s['ot_h'] += meta.get("overtime_hours", 0.0)
            if is_leave:         s['leave']  += 1
            if '曠職' in remark: s['absent'] += 1
            if '遲到' in remark: s['late']   += 1
            if '早退' in remark: s['early']  += 1
        return [
            (eid, d['name'], d['sched'], d['work'],
             round(d['total_h'], 1),
             round(d['ot_h'], 1) if d['ot_h'] else "",
             d['leave'], d['absent'], d['late'], d['early'])
            for eid, d in sorted(stats.items())
        ]

    def _show_attendance_stats(self) -> None:
        rows = self._compute_attendance_stats()
        if not rows:
            messagebox.showinfo("提示", "請先查詢打卡資料。")
            return
        start = self.p_start.get().strip()
        end   = self.p_end.get().strip()

        top = ctk.CTkToplevel(self)
        top.title(f"出勤統計　{start}～{end}")
        top.geometry("920x520")
        top.configure(fg_color=_C["app_bg"])
        top.transient(self)

        _COLS = [
            ("員工代號", 80), ("員工姓名", 90), ("應出勤", 60), ("實出勤", 60),
            ("工時(h)", 65), ("加班(h)", 65), ("請假(天)", 65),
            ("曠職", 50), ("遲到", 50), ("早退", 50),
        ]
        col_ids = [f"c{i}" for i in range(len(_COLS))]

        wrap = tk.Frame(top, bg=_C["app_bg"])
        wrap.pack(fill="both", expand=True, padx=10, pady=(10, 0))

        tv = ttk.Treeview(wrap, style="App.Treeview",
                          columns=col_ids, show="headings", height=18)
        for cid, (hdr, w) in zip(col_ids, _COLS):
            tv.heading(cid, text=hdr)
            anc = "w" if hdr in ("員工代號", "員工姓名") else "center"
            tv.column(cid, width=w, anchor=anc, stretch=False)
        for row in rows:
            tv.insert("", "end", values=row)
        vsb = ttk.Scrollbar(wrap, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        tv.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        _bind_mousewheel(tv)

        def _export():
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 活頁簿", "*.xlsx")],
                initialfile=f"出勤統計_{start}_{end}.xlsx",
                parent=top,
            )
            if not path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "出勤統計"
            hdrs = [h for h, _ in _COLS]
            ws.append(hdrs)
            hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                   fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = hdr_fill
            for row in rows:
                ws.append(list(row))
            col_ws = [10, 12, 8, 8, 9, 9, 9, 7, 7, 7]
            for i, w in enumerate(col_ws, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            for r in ws.iter_rows(min_row=2):
                for cell in r:
                    cell.alignment = Alignment(horizontal="center")
                r[1].alignment = Alignment(horizontal="left")
            try:
                wb.save(path)
                messagebox.showinfo("完成", f"已匯出 {len(rows)} 人統計 → {path}",
                                    parent=top)
            except Exception as exc:
                messagebox.showerror("匯出失敗", str(exc), parent=top)

        ctk.CTkButton(top, text="匯出 Excel", width=120,
                       fg_color=_C["muted"], hover_color="#6B7280",
                       corner_radius=8, command=_export).pack(pady=8)

    # -------------------------------------------------------------- 特休餘額
    def _show_annual_leave(self) -> None:
        """彈出視窗顯示目前排班篩選下員工的特休剩餘天數與到期日期。"""
        if not self._need_client():
            return
        if not self.schedule:
            messagebox.showinfo("尚未載入", "請先按「載入排班」載入員工後再查特休。")
            return

        emps = []
        for emp in self.schedule:
            if not self._emp_passes_filter(emp.emp_id, emp):
                continue
            raw = emp.raw or {}
            dept_code = (
                _first_field(raw, _DEPT_CODE_KEYS)
                or self._emp_dept_code.get(emp.emp_id, "")
                or self._dept_code_key(self._schedule_emp_dept(emp))
                or self._dept_code_key(self._emp_filter_dept(emp.emp_id))
            )
            emps.append({
                "emp_id": emp.emp_id,
                "name": emp.name,
                "dept": dept_code,
            })

        if not emps:
            messagebox.showinfo("無員工", "目前篩選後沒有員工可查。")
            return

        self._set_status("查詢特休中…")
        self._show_progress(0, len(emps))

        def work():
            def on_progress(done_n: int, total: int) -> None:
                self.after(0, lambda d=done_n, t=total: (
                    self._show_progress(d, t),
                    self._set_status(f"查詢特休中  {d} / {t}"),
                ))
            return self.client.get_annual_leave_balance_bulk(
                emps, progress_cb=on_progress
            )

        def done(rows):
            self._hide_progress()
            self._set_status(f"特休查詢完成:{len(emps)} 位員工")
            self._render_annual_leave_window(rows)

        self._run_async(work, done)

    def _render_annual_leave_window(self, rows: list) -> None:
        today = dt.date.today()

        def _fmt_date(value) -> str:
            text = str(value or "").strip()
            return text[:10] if len(text) >= 10 else text

        def _soon(end_str, left_days) -> bool:
            try:
                ed = dt.date.fromisoformat(_fmt_date(end_str))
            except ValueError:
                return False
            try:
                ld = float(left_days)
            except (TypeError, ValueError):
                ld = 0.0
            return ld > 0 and today <= ed <= today + dt.timedelta(days=90)

        cols = [
            ("員工代號", 80), ("員工姓名", 90), ("年度", 55), ("年資", 70),
            ("剩餘天數", 75), ("剩餘時數", 75), ("給假天數", 75), ("已用天數", 75),
            ("啟用日期", 95), ("到期日期", 95), ("備註", 130),
        ]
        col_ids = [f"c{i}" for i in range(len(cols))]

        disp = []
        for row in rows:
            note = row.get("note") or row.get("error") or ""
            left_min = row.get("left_minutes") or 0
            left_h = round(left_min / 60, 1) if left_min else 0
            year = row.get("year")
            seniority = row.get("seniority")
            vals = (
                row.get("emp_id", ""),
                row.get("name", ""),
                "" if year in (None, "") else year,
                "" if seniority in (None, "") else seniority,
                "" if note else row.get("left_days", 0),
                "" if note else left_h,
                "" if note else row.get("granted_days", 0),
                "" if note else row.get("used_days", 0),
                _fmt_date(row.get("start_date", "")),
                _fmt_date(row.get("end_date", "")),
                note,
            )
            if row.get("error"):
                tag = "err"
            elif _soon(row.get("end_date", ""), row.get("left_days")):
                tag = "soon"
            else:
                tag = ""
            disp.append((vals, tag))

        top = ctk.CTkToplevel(self)
        top.title(f"特休餘額　共 {len(disp)} 列　查詢日 {today.isoformat()}")
        top.geometry("1040x560")
        top.configure(fg_color=_C["app_bg"])
        top.transient(self)

        ctk.CTkLabel(
            top,
            text="天數以每日 8 小時換算。底色:橘=90 天內到期且尚有餘額,紅=查詢失敗。",
            text_color=_C["muted"], font=("", 12),
        ).pack(anchor="w", padx=12, pady=(10, 0))

        wrap = tk.Frame(top, bg=_C["app_bg"])
        wrap.pack(fill="both", expand=True, padx=10, pady=(6, 0))

        tv = ttk.Treeview(wrap, style="App.Treeview",
                          columns=col_ids, show="headings", height=18)
        for cid, (hdr, width) in zip(col_ids, cols):
            tv.heading(cid, text=hdr)
            anchor = "w" if hdr in ("員工代號", "員工姓名", "備註") else "center"
            tv.column(cid, width=width, anchor=anchor, stretch=False)
        tv.tag_configure("soon", background="#FFE8D6")
        tv.tag_configure("err", background="#FBE0DC")
        for vals, tag in disp:
            tv.insert("", "end", values=vals, tags=(tag,) if tag else ())
        vsb = ttk.Scrollbar(wrap, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        tv.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        _bind_mousewheel(tv)

        def _export():
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 活頁簿", "*.xlsx")],
                initialfile=f"特休餘額_{today.isoformat()}.xlsx",
                parent=top,
            )
            if not path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "特休餘額"
            ws.append([h for h, _ in cols])
            hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                   fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = hdr_fill
            soon_fill = PatternFill(start_color="FFE8D6", end_color="FFE8D6",
                                    fill_type="solid")
            err_fill = PatternFill(start_color="FBE0DC", end_color="FBE0DC",
                                   fill_type="solid")
            for vals, tag in disp:
                ws.append(list(vals))
                fill = soon_fill if tag == "soon" else (
                    err_fill if tag == "err" else None)
                if fill:
                    for cell in ws[ws.max_row]:
                        cell.fill = fill
            col_ws = [10, 12, 7, 10, 9, 9, 9, 9, 12, 12, 18]
            for i, width in enumerate(col_ws, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                row[1].alignment = Alignment(horizontal="left")
                row[10].alignment = Alignment(horizontal="left")
            try:
                wb.save(path)
                messagebox.showinfo("完成", f"已匯出 {len(disp)} 列特休 → {path}",
                                    parent=top)
            except Exception as exc:
                messagebox.showerror("匯出失敗", str(exc), parent=top)

        ctk.CTkButton(top, text="匯出 Excel", width=120,
                       fg_color=_C["muted"], hover_color="#6B7280",
                       corner_radius=8, command=_export).pack(pady=8)

    # -------------------------------------------------------------- 到職日
    def _show_staff_info(self) -> None:
        """顯示目前排班篩選下員工的到職日與年資。"""
        if not self.schedule:
            messagebox.showinfo("尚未載入", "請先按「載入排班」載入員工後再查到職日。")
            return

        today = dt.date.today()

        def _fmt_date(value) -> str:
            text = str(value or "").strip()
            return text[:10] if len(text) >= 10 else text

        def _seniority(hire_iso: str) -> str:
            try:
                hire_date = dt.date.fromisoformat(_fmt_date(hire_iso))
            except ValueError:
                return ""
            months = (today.year - hire_date.year) * 12 + (today.month - hire_date.month)
            if today.day < hire_date.day:
                months -= 1
            if months < 0:
                return ""
            return f"{months // 12}年{months % 12}月"

        rows = []
        for emp in self.schedule:
            if not self._emp_passes_filter(emp.emp_id, emp):
                continue
            hire = _fmt_date(getattr(emp, "hire_date", "") or emp.raw.get("pa51024", ""))
            rows.append((emp.emp_id, emp.name, emp.title or "", hire, _seniority(hire)))
        if not rows:
            messagebox.showinfo("無員工", "目前篩選後沒有員工可顯示。")
            return
        rows.sort(key=lambda row: (row[3] == "", row[3]))

        cols = [("員工代號", 90), ("員工姓名", 110), ("職稱", 150),
                ("到職日", 110), ("年資", 90)]
        col_ids = [f"c{i}" for i in range(len(cols))]

        top = ctk.CTkToplevel(self)
        top.title(f"夥伴到職日　共 {len(rows)} 人")
        top.geometry("620x520")
        top.configure(fg_color=_C["app_bg"])
        top.transient(self)

        ctk.CTkLabel(
            top, text="依到職日排序（資深在前）。年資為到今日的概算。",
            text_color=_C["muted"], font=("", 12),
        ).pack(anchor="w", padx=12, pady=(10, 0))

        wrap = tk.Frame(top, bg=_C["app_bg"])
        wrap.pack(fill="both", expand=True, padx=10, pady=(6, 0))
        tv = ttk.Treeview(wrap, style="App.Treeview",
                          columns=col_ids, show="headings", height=18)
        for cid, (hdr, width) in zip(col_ids, cols):
            tv.heading(cid, text=hdr)
            anchor = "center" if hdr in ("到職日", "年資") else "w"
            tv.column(cid, width=width, anchor=anchor, stretch=False)
        for row in rows:
            tv.insert("", "end", values=row)
        vsb = ttk.Scrollbar(wrap, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        tv.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        _bind_mousewheel(tv)

        def _export():
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 活頁簿", "*.xlsx")],
                initialfile=f"夥伴到職日_{today.isoformat()}.xlsx",
                parent=top,
            )
            if not path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "到職日"
            ws.append([h for h, _ in cols])
            hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                   fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = hdr_fill
            for row in rows:
                ws.append(list(row))
            col_ws = [12, 14, 20, 14, 10]
            for i, width in enumerate(col_ws, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                for idx in (0, 1, 2):
                    row[idx].alignment = Alignment(horizontal="left")
            try:
                wb.save(path)
                messagebox.showinfo("完成", f"已匯出 {len(rows)} 人到職日 → {path}",
                                    parent=top)
            except Exception as exc:
                messagebox.showerror("匯出失敗", str(exc), parent=top)

        ctk.CTkButton(top, text="匯出 Excel", width=120,
                       fg_color=_C["muted"], hover_color="#6B7280",
                       corner_radius=8, command=_export).pack(pady=8)

    def _import_schedule(self) -> None:
        if not self._need_client():
            return
        path = filedialog.askopenfilename(
            filetypes=[("Excel 活頁簿", "*.xlsx *.xls")],
            title="選擇排班 Excel",
        )
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:
            messagebox.showerror("讀取失敗", str(exc))
            return
        if len(rows) < 2:
            messagebox.showwarning("空檔", "Excel 沒有資料列。")
            return
        # 第一列找日期欄（YYYY-MM-DD 格式）
        header = rows[0]
        date_cols: dict[int, str] = {}
        for i, h in enumerate(header):
            h_str = str(h).strip() if h is not None else ""
            if len(h_str) == 10 and h_str[4] == "-" and h_str[7] == "-":
                try:
                    dt.date.fromisoformat(h_str)
                    date_cols[i] = h_str
                except ValueError:
                    pass
        if not date_cols:
            messagebox.showerror(
                "格式錯誤",
                "找不到日期欄（YYYY-MM-DD）。\n請使用本工具匯出的 Excel 再匯入。",
            )
            return
        first_date = next(iter(date_cols.values()))
        year, month = int(first_date[:4]), int(first_date[5:7])
        # 收集有班別的格子，同一 emp+date 取最後一筆
        seen: dict[tuple, dict] = {}
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            emp_id = str(row[0]).strip()
            for col_idx, date_str in date_cols.items():
                raw = row[col_idx] if col_idx < len(row) else None
                code = str(raw).strip() if raw is not None else ""
                if code:
                    seen[(emp_id, date_str)] = {
                        "emp_id": emp_id, "date": date_str, "shift_code": code
                    }
        changes = list(seen.values())
        if not changes:
            messagebox.showinfo("無班別", "沒有找到任何班別資料。")
            return

        # 套用排班頁部門 / 員工篩選（若有設定）
        if self._selected_depts or self._selected_emps:
            changes = [ch for ch in changes
                       if self._emp_passes_filter(ch["emp_id"])]
            if not changes:
                messagebox.showinfo(
                    "無符合員工",
                    "目前排班篩選條件在這份 Excel 中沒有班別資料。")
                return
        n_emps = len({ch["emp_id"] for ch in changes})
        filt_note = "（已套用排班篩選）" if (
            self._selected_depts or self._selected_emps
        ) else ""
        if not messagebox.askyesno(
            "確認匯入",
            f"將匯入 {year}-{month:02d} 排班，共 {len(changes)} 筆"
            f"（{n_emps} 位員工）{filt_note}。\n\n"
            "這會寫入正式班表，確定？",
        ):
            return
        self._set_status(f"匯入中…")
        self._show_progress(0, len(changes))

        def work():
            for ch in changes:
                if _is_rest(ch["shift_code"]):
                    ch.setdefault("kind", 3)   # 排休
                else:
                    ch.setdefault("kind", 1)   # 工作日

            def on_progress(done_n: int, total: int) -> None:
                self.after(0, lambda d=done_n, t=total: (
                    self._show_progress(d, t),
                    self._set_status(f"匯入中  {d} / {t}"),
                ))

            return self.client.set_shifts_bulk(
                year, month, changes, dry_run=False, progress_cb=on_progress
            )

        def done(results):
            self._hide_progress()
            ok_n = sum(1 for r in results if r.get("ok"))
            fail_n = len(results) - ok_n
            self._set_status(f"匯入完成：{ok_n} 筆成功，{fail_n} 筆失敗")
            if fail_n:
                errs = [
                    f"{r['emp_id']} {r['date']}：{r.get('error', '未知')}"
                    for r in results if not r.get("ok")
                ]
                messagebox.showwarning(
                    "部分失敗",
                    f"{fail_n} 筆寫入失敗：\n" + "\n".join(errs[:10])
                    + ("\n…" if fail_n > 10 else ""),
                )
            self._load_schedule()

        self._run_async(work, done)

    def _pick_employees(
        self,
        emps: dict[str, str],
        preselect: set[str] | None = None,
        *,
        title: str = "選擇員工",
        hint: str = "勾選要處理的員工",
        ok_text: str = "確定",
    ) -> set[str] | None:
        """通用員工挑選對話框（主畫面篩選與匯入共用）。
        preselect=None 代表預設全選；否則只勾選集合內的員工。
        回傳勾選的 emp_id 集合，使用者關閉視窗回傳 None。"""
        top = ctk.CTkToplevel(self)
        top.title(title)
        top.geometry("300x520")
        top.configure(fg_color=_C["app_bg"])
        top.resizable(False, True)
        top.transient(self)
        top.after(60, top.grab_set)

        result: dict = {"value": None}

        ctk.CTkLabel(top, text=hint,
                      text_color=_C["muted"], font=("", 12)).pack(
            padx=12, pady=(12, 4))

        search_var = ctk.StringVar()
        ctk.CTkEntry(top, placeholder_text="搜尋員工…",
                      textvariable=search_var,
                      border_color=_C["line"], fg_color=_C["white"],
                      text_color=_C["ink"]).pack(fill="x", padx=12, pady=(0, 4))

        scroll = ctk.CTkScrollableFrame(top, fg_color=_C["white"],
                                          corner_radius=8, border_width=1,
                                          border_color=_C["line"])
        _bind_scrollable_frame_mousewheel(scroll)
        scroll.pack(fill="both", expand=True, padx=12, pady=(4, 0))

        chk_vars: dict[str, tk.BooleanVar] = {}
        chk_widgets: dict[str, ctk.CTkCheckBox] = {}
        for eid in sorted(emps):
            checked = True if preselect is None else (eid in preselect)
            var = tk.BooleanVar(value=checked)
            label = f"{eid}  {emps[eid]}".rstrip()
            chk = ctk.CTkCheckBox(scroll, text=label,
                                   variable=var, text_color=_C["ink"])
            chk.pack(anchor="w", padx=4, pady=3)
            chk_vars[eid] = var
            chk_widgets[eid] = chk

        def _filter(*_):
            q = search_var.get().lower()
            for eid, w in chk_widgets.items():
                if not q or q in eid.lower() or q in emps[eid].lower():
                    w.pack(anchor="w", padx=4, pady=2)
                else:
                    w.pack_forget()
        search_var.trace_add("write", _filter)

        btn_bar = ctk.CTkFrame(top, fg_color="transparent")
        btn_bar.pack(fill="x", padx=12, pady=(6, 0))
        ctk.CTkButton(btn_bar, text="全選", width=80,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"], corner_radius=8,
                       command=lambda: [v.set(True) for v in chk_vars.values()]
                       ).pack(side="left", padx=(0, 4))
        ctk.CTkButton(btn_bar, text="清除", width=80,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"], corner_radius=8,
                       command=lambda: [v.set(False) for v in chk_vars.values()]
                       ).pack(side="left")

        def apply():
            result["value"] = {eid for eid, v in chk_vars.items() if v.get()}
            top.destroy()

        ctk.CTkButton(top, text=ok_text,
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=9, font=("", 13, "bold"),
                       command=apply).pack(fill="x", padx=12, pady=10)

        top.wait_window()
        return result["value"]

    # ── 排班頁部門 / 員工篩選 ───────────────────────────────────────────────
    def _fetch_emp_dept_map(self, start: dt.date, end: dt.date) -> dict[str, str]:
        rows = None
        # B1:若剛在打卡分頁查過同一區間,直接重用那批資料(含部門欄位),免重抓。
        if (self._raw_punch_range == (start.isoformat(), end.isoformat())
                and self._raw_punch_rows):
            rows = self._raw_punch_rows
        if rows is None:
            try:
                rows = self.client.get_punch_records(start, end, readable=False)
            except Exception:
                return {}
        dept_map: dict[str, str] = {}
        base_dept = ""
        for row in rows:
            eid = _first_field(row, _EMP_ID_KEYS)
            dept = self._row_dept(row)
            if eid and dept:
                dept_map[eid] = dept
                name = self._dept_name_key(dept)
                if not base_dept and name and "維修中心" not in name and "(前)" not in name:
                    base_dept = dept
        if base_dept and (end - start).days == 27 and start.weekday() == 0:
            for eid in self._fetch_period_emp_ids(start, end):
                dept_map.setdefault(eid, base_dept)
        return dept_map

    def _fetch_period_emp_ids(self, start: dt.date, end: dt.date) -> set[str]:
        emp_ids: set[str] = set()
        for flex_type in (0, 1, 2):
            try:
                cal = self.client.get_schedule_raw(
                    start.year, start.month,
                    shiftType=1, flexType=flex_type,
                    start=f"{start.isoformat()}T00:00:00.000",
                    end=f"{end.isoformat()}T23:59:59.000",
                )
            except Exception:
                continue
            for emp in cal.get("shiftEmployees", []):
                eid = str(emp.get("pa51002") or "").strip()
                if eid:
                    emp_ids.add(eid)
        return emp_ids

    def _schedule_emp_dept(self, emp) -> str:
        raw = getattr(emp, "raw", {}) or {}
        emp_id = getattr(emp, "emp_id", "")
        return self._emp_filter_dept(emp_id, row=raw)

    def _refresh_sched_emp_dept(self, sched) -> None:
        self._sched_emp_dept = {}
        for emp in sched:
            dept = self._schedule_emp_dept(emp)
            if dept:
                self._sched_emp_dept[emp.emp_id] = dept

    def _emp_passes_filter(self, emp_id: str, emp=None) -> bool:
        """空篩選集合代表「全部」；否則同時套用部門與員工條件。"""
        dept = self._sched_emp_dept.get(emp_id, "")
        if not dept and emp is not None:
            dept = self._schedule_emp_dept(emp)
            if dept:
                self._sched_emp_dept[emp_id] = dept
        if not dept:
            dept = self._emp_filter_dept(emp_id)
        if not self._dept_passes_filter(dept):
            return False
        if self._selected_emps and emp_id not in self._selected_emps:
            return False
        return True

    def _update_sched_filter_btns(self) -> None:
        self._update_filter_buttons()

    def _refresh_schedule_view(self) -> None:
        if not self.schedule:
            return
        if self._grid_dates:
            self._populate_grid_period(self._grid_dates, self.schedule)
            return
        try:
            self._populate_grid(int(self.year_var.get()), int(self.month_var.get()),
                                self.schedule)
        except ValueError:
            pass

    def _update_filter_buttons(self) -> None:
        dept_n = len(self._selected_depts)
        emp_n  = len(self._selected_emps)
        dept_text = "部門：全部 ▼" if dept_n == 0 else f"部門：已選 {dept_n} ▼"
        emp_text  = "員工：全員 ▼" if emp_n == 0 else f"員工：已選 {emp_n} 人 ▼"
        for attr in ("sched_dept_btn", "p_dept_btn", "raw_dept_btn", "suggest_dept_btn"):
            if hasattr(self, attr):
                getattr(self, attr).configure(text=dept_text)
        for attr in ("sched_emp_btn", "p_emp_btn", "raw_emp_btn", "suggest_emp_btn"):
            if hasattr(self, attr):
                getattr(self, attr).configure(text=emp_text)

    def _row_dept_code(self, row: dict | None) -> str:
        return _first_field(row or {}, _DEPT_CODE_KEYS)

    def _row_dept(self, row: dict | None) -> str:
        row = row or {}
        full = _first_field(row, _DEPT_FULL_KEYS)
        if full:
            return full
        code = self._row_dept_code(row)
        name = _first_field(row, _DEPT_NAME_KEYS)
        if code and name:
            return f"{code} {name}"
        return name or code or _first_field(row, _DEPT_KEYS)

    def _dept_code_key(self, dept: str) -> str:
        match = re.match(r"^\s*(\d+(?:-\d+)?)\b", str(dept or ""))
        return match.group(1) if match else ""

    def _dept_name_key(self, dept: str) -> str:
        dept = str(dept or "").strip()
        return re.sub(r"^\d+\s*[-－_:：]?\s*", "", dept).strip()

    def _dept_matches_one(self, dept: str, selected: str) -> bool:
        dept_code = self._dept_code_key(dept)
        selected_code = self._dept_code_key(selected)
        if dept_code and selected_code:
            return dept_code == selected_code
        dept_key = self._dept_name_key(dept)
        selected_key = self._dept_name_key(selected)
        return bool(dept_key and selected_key and dept_key == selected_key)

    def _dept_passes_filter(self, dept: str) -> bool:
        if not self._selected_depts:
            return True
        return any(self._dept_matches_one(dept, selected)
                   for selected in self._selected_depts)

    def _remember_emp_dept(self, emp_id: str, dept: str) -> None:
        emp_id = str(emp_id or "").strip()
        dept = str(dept or "").strip()
        if emp_id and dept:
            self._emp_dept[emp_id] = dept
            code = self._dept_code_key(dept)
            if code:
                self._emp_dept_code[emp_id] = code

    def _emp_filter_dept(self, emp_id: str, row: dict | None = None) -> str:
        dept = self._row_dept(row)
        if dept:
            self._remember_emp_dept(emp_id, dept)
            return dept
        return (
            self._emp_dept.get(emp_id, "")
            or getattr(self, "_sched_emp_dept", {}).get(emp_id, "")
        )

    def _prune_selected_emps_for_depts(self) -> None:
        if not self._selected_depts or not self._selected_emps:
            return
        self._selected_emps = {
            eid for eid in self._selected_emps
            if self._dept_passes_filter(self._emp_filter_dept(eid))
        }

    def _set_global_filter(
        self,
        *,
        depts: set[str] | None = None,
        emps: set[str] | None = None,
    ) -> None:
        if depts is not None:
            self._selected_depts = set(depts)
            self._prune_selected_emps_for_depts()
        if emps is not None:
            self._selected_emps = set(emps)
        self._update_filter_buttons()
        self._refresh_schedule_view()
        if hasattr(self, "punch_tbl"):
            self._apply_punch_filter()
        elif hasattr(self, "suggest_table"):
            self._compute_suggestions()
        if hasattr(self, "raw_tv"):
            self._populate_raw_punch()

    def _open_sched_dept_picker(self) -> None:
        if not self.schedule:
            messagebox.showinfo("尚未載入", "請先按「載入排班」載入員工後再篩選。")
            return
        depts = sorted(set(self._sched_emp_dept.values()) - {""})
        if not depts:
            messagebox.showinfo("沒有部門資料", "目前排班資料沒有可篩選的部門欄位。請重新載入排班後再試。")
            return
        self._generic_picker(
            title="選擇部門",
            items=depts,
            selected=self._selected_depts,
            on_confirm=lambda s: self._set_global_filter(depts=s)
        )

    def _open_sched_emp_picker(self) -> None:
        if not self.schedule:
            messagebox.showinfo("尚未載入", "請先按「載入排班」載入員工後再篩選。")
            return
        emps: dict[str, str] = {}
        for emp in self.schedule:
            dept = self._schedule_emp_dept(emp)
            if not self._dept_passes_filter(dept):
                continue
            emps[emp.emp_id] = emp.name
        if not emps:
            messagebox.showinfo("無符合員工", "目前部門篩選下沒有可選員工。")
            return
        items = [f"{eid}  {name}".rstrip() for eid, name in sorted(emps.items())]
        selected = {
            f"{eid}  {emps[eid]}".rstrip()
            for eid in self._selected_emps
            if eid in emps
        }
        self._generic_picker(
            title="選擇員工",
            items=items,
            selected=selected,
            on_confirm=lambda s: self._set_global_filter(
                emps={item.split()[0] for item in s}
            ),
            searchable=True,
        )

    # ---------------------------------------------------------------- punch
    def _query_punch(self) -> None:
        if not self._need_client():
            return
        start = self.p_start.get().strip()
        end = self.p_end.get().strip()
        self._set_status("查詢打卡中…")

        def work():
            punch_rows = self.client.get_punch_records(
                start, end, readable=True
            )
            try:
                leave_rows = self.client.get_leave_records(start, end)
            except Exception:
                leave_rows = []
            start_d = dt.date.fromisoformat(start)
            end_d   = dt.date.fromisoformat(end)
            # 收集區間橫跨的所有年月,並行抓排班(取代逐月序列迴圈)
            months: set[tuple[int, int]] = set()
            cur = start_d.replace(day=1)
            while cur <= end_d:
                months.add((cur.year, cur.month))
                cur = (cur.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
            results, failures = self._fetch_month_schedules(months)
            sched_map: dict[tuple, tuple] = {}
            emp_names: dict[str, str] = {}
            for ym in sorted(results):
                for es in results[ym]:
                    emp_names[es.emp_id] = es.name
                    for s in es.shifts:
                        sched_map[(es.emp_id, s.date)] = (s.code, s.name, s.kind)
            return punch_rows, leave_rows, sched_map, emp_names, start, end, failures

        def done(result):
            punch_rows, leave_rows, sched_map, emp_names, s, e, failures = result
            self._raw_punch_rows = list(punch_rows)
            self._raw_punch_range = (s, e)
            self._populate_punch(punch_rows, sched_map, emp_names, s, e,
                                 leave_rows=leave_rows)
            self._populate_raw_punch()
            self._set_status(f"打卡 {start}~{end} 共 {len(punch_rows)} 筆"
                             + self._fmt_month_failures(failures))

        self._run_async(work, done)

    def _populate_punch(
        self,
        rows,
        sched_map: dict | None = None,
        emp_names: dict | None = None,
        range_start: str | None = None,
        range_end: str | None = None,
        leave_rows: list | None = None,
    ) -> None:
        # 0. 建立 emp_id -> 部門名稱 對照表（從原始打卡資料）
        for r in rows:
            eid = _first_field(r, _EMP_ID_KEYS)
            self._remember_emp_dept(eid, self._row_dept(r))
        if self.schedule:
            self._refresh_sched_emp_dept(self.schedule)

        # 1. 整理有打卡的日子
        sorted_rows = sorted(rows, key=lambda r: (
            r.get("員工代號", ""),
            str(r.get("出勤日期", ""))[:10],
            str(r.get("刷卡時間", "")),
        ))
        display: list[tuple] = []
        punched: set[tuple] = set()
        for (emp_id, date), grp in groupby(sorted_rows, key=lambda r: (
            r.get("員工代號", ""), str(r.get("出勤日期", ""))[:10]
        )):
            punched.add((emp_id, date))
            group = list(grp)
            name = group[0].get("員工姓名", "")
            _sm = (sched_map or {}).get((emp_id, date), ("", "", None))
            code, _, s_kind = (_sm + (None,))[:3]
            sched_in, sched_out = _shift_times(code) if code else ("", "")
            # 提取打卡時間，區分有無結轉標記
            all_punches = [(str(r.get("刷卡時間", ""))[11:16],
                            r.get("結轉註記", "")) for r in group]
            jiezhuan = [t for t, flag in all_punches if flag == "結轉" and t]
            all_times  = [t for t, _  in all_punches if t]
            # 有結轉打卡 → 優先使用，用寬鬆分群；否則用全部打卡與原分群邏輯
            if jiezhuan:
                use_punches = jiezhuan
                split_fn = _split_punches_jiezhuan
            else:
                use_punches = all_times
                split_fn = _split_punches
            if len(use_punches) == 1:
                clock_in, clock_out = _assign_single_punch(use_punches[0], sched_in, sched_out)
            else:
                clock_in, clock_out = split_fn(use_punches, sched_in, sched_out)
            remark = _punch_remark(clock_in, clock_out, sched_in, sched_out, kind=s_kind)
            vals = (emp_id, name, date, code, sched_in, sched_out, clock_in, clock_out, remark)
            meta = {"category": _classify_remark(remark),
                    "overtime_hours": _overtime_hours(clock_in, clock_out, sched_in, sched_out, kind=s_kind)}
            display.append((vals, meta))

        # 1b. 建立 leave_map：(emp_id, date) → 假別名稱（特休/病假/事假…）
        leave_map: dict[tuple, str] = {}
        for lr in (leave_rows or []):
            eid   = lr.get("pa60002", "")
            dstr  = str(lr.get("pa60006", ""))[:10]
            lname = lr.get("pa60004Name", "")
            if eid and dstr and lname:
                key = (eid, dstr)
                if key not in leave_map:
                    leave_map[key] = lname
                elif lname not in leave_map[key]:
                    leave_map[key] += f"/{lname}"

        # 2. 找出有排班但完全沒打卡的日子 → 查假單；無假單才標「曠職」
        if sched_map and range_start and range_end:
            start_d = dt.date.fromisoformat(range_start)
            end_d   = dt.date.fromisoformat(range_end)
            for (emp_id, date), sched_val in sched_map.items():
                code = sched_val[0]
                if not code or _is_rest(code):
                    continue
                d = dt.date.fromisoformat(date)
                if start_d <= d <= end_d and (emp_id, date) not in punched:
                    name = (emp_names or {}).get(emp_id, "")
                    sched_in, sched_out = _shift_times(code) if code else ("", "")
                    remark = leave_map.get((emp_id, date), "曠職")
                    vals = (emp_id, name, date, code, sched_in, sched_out, "", "", remark)
                    display.append((vals, {"category": _classify_remark(remark),
                                           "overtime_hours": 0.0}))

        RED    = "#CC0000"
        GREEN  = "#2A8B5A"
        ORANGE = "#C07000"
        def _fgs(row: tuple) -> tuple:
            remark = row[8]
            cat = _classify_remark(remark)
            ci_red = RED if remark and any(k in remark for k in ("遲到", "上班忘打卡", "曠職")) else ""
            co_red = RED if remark and any(k in remark for k in ("早退", "下班忘打卡", "曠職")) else ""
            rm_col = RED if cat == "err" else (ORANGE if cat == "ot" else (GREEN if cat == "leave" else ""))
            return ("", "", "", "", "", "", ci_red, co_red, rm_col)

        # 每列存 (顯示值, 著色, meta);meta 含 category / overtime_hours,供統計與篩選直接讀。
        self._punch_cache = [
            (vals, _fgs(vals), meta)
            for vals, meta in sorted(display, key=lambda x: (x[0][0], x[0][2]))
        ]
        self._apply_punch_filter()

    def _apply_punch_filter(self) -> None:
        show_err   = self.p_show_err.get()
        show_ot    = self.p_show_ot.get()
        show_leave = self.p_show_leave.get()
        any_filter = show_err or show_ot or show_leave

        def _cat_match(vals) -> bool:
            if not any_filter:
                return True
            remark = vals[8] or ""
            is_err   = any(k in remark for k in _REMARK_ERR_KEYWORDS)
            is_ot    = _REMARK_OT_KEYWORD in remark
            is_leave = bool(remark) and not is_err and not is_ot
            return (show_err and is_err) or (show_ot and is_ot) or (show_leave and is_leave)

        filtered = [
            (vals, fgs) for vals, fgs, _meta in self._punch_cache
            if self._dept_passes_filter(self._emp_filter_dept(vals[0]))
            and (not self._selected_emps or vals[0] in self._selected_emps)
            and _cat_match(vals)
        ]
        self.punch_tbl.populate(filtered)
        self.after(50, self._compute_suggestions)

    def _update_dept_btn(self) -> None:
        self._update_filter_buttons()

    def _update_emp_btn(self) -> None:
        self._update_filter_buttons()

    def _open_dept_picker(self) -> None:
        # 從 _emp_dept 收集所有不重複部門
        depts = sorted(
            (set(self._emp_dept.values()) | set(getattr(self, "_sched_emp_dept", {}).values())) - {""}
        )
        if not depts:
            messagebox.showinfo("提示", "請先查詢打卡資料。")
            return

        top = ctk.CTkToplevel(self)
        top.title("選擇部門")
        top.geometry("260x400")
        top.configure(fg_color=_C["app_bg"])
        top.resizable(False, True)
        top.transient(self)
        top.after(60, top.grab_set)

        scroll = ctk.CTkScrollableFrame(top, fg_color=_C["white"],
                                          corner_radius=8,
                                          border_width=1, border_color=_C["line"])
        _bind_scrollable_frame_mousewheel(scroll)
        scroll.pack(fill="both", expand=True, padx=12, pady=(12, 0))

        chk_vars: dict[str, tk.BooleanVar] = {}
        for dept in depts:
            init = self._dept_passes_filter(dept)
            var = tk.BooleanVar(value=init)
            ctk.CTkCheckBox(scroll, text=dept, variable=var,
                             text_color=_C["ink"]).pack(anchor="w", padx=4, pady=3)
            chk_vars[dept] = var

        btn_bar = ctk.CTkFrame(top, fg_color="transparent")
        btn_bar.pack(fill="x", padx=12, pady=(6, 0))
        ctk.CTkButton(btn_bar, text="全選", width=80,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"], corner_radius=8,
                       command=lambda: [v.set(True) for v in chk_vars.values()]
                       ).pack(side="left", padx=(0, 4))
        ctk.CTkButton(btn_bar, text="清除", width=80,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"], corner_radius=8,
                       command=lambda: [v.set(False) for v in chk_vars.values()]
                       ).pack(side="left")

        def apply():
            selected = {d for d, v in chk_vars.items() if v.get()}
            self._set_global_filter(
                depts=set() if len(selected) == len(depts) else selected
            )
            top.destroy()

        ctk.CTkButton(top, text="確定",
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=9, font=("", 13, "bold"),
                       command=apply).pack(fill="x", padx=12, pady=10)

    def _open_emp_picker(self) -> None:
        # 從打卡快取收集所有員工
        emps: dict[str, str] = {}
        for vals, *_ in self._punch_cache:
            dept = self._emp_filter_dept(vals[0])
            if not self._dept_passes_filter(dept):
                continue
            emps[vals[0]] = vals[1]

        if not emps:
            messagebox.showinfo("提示", "請先查詢打卡資料。")
            return

        top = ctk.CTkToplevel(self)
        top.title("選擇員工")
        top.geometry("280x490")
        top.configure(fg_color=_C["app_bg"])
        top.resizable(False, True)
        top.transient(self)
        top.after(60, top.grab_set)

        # 搜尋框
        search_var = ctk.StringVar()
        ctk.CTkEntry(top, placeholder_text="搜尋員工…",
                      textvariable=search_var,
                      border_color=_C["line"], fg_color=_C["white"],
                      text_color=_C["ink"]).pack(fill="x", padx=12, pady=(12, 4))

        # 捲動式 checkbox 清單
        scroll = ctk.CTkScrollableFrame(top, fg_color=_C["white"],
                                          corner_radius=8,
                                          border_width=1, border_color=_C["line"])
        _bind_scrollable_frame_mousewheel(scroll)
        scroll.pack(fill="both", expand=True, padx=12, pady=(4, 0))

        chk_vars: dict[str, tk.BooleanVar] = {}
        chk_widgets: dict[str, ctk.CTkCheckBox] = {}
        for eid in sorted(emps):
            init = (not self._selected_emps) or (eid in self._selected_emps)
            var = tk.BooleanVar(value=init)
            chk = ctk.CTkCheckBox(scroll, text=f"{eid}  {emps[eid]}",
                                   variable=var, text_color=_C["ink"])
            chk.pack(anchor="w", padx=4, pady=3)
            chk_vars[eid] = var
            chk_widgets[eid] = chk

        def _filter(*_):
            q = search_var.get().lower()
            for eid, w in chk_widgets.items():
                show = not q or q in eid.lower() or q in emps[eid].lower()
                if show:
                    w.pack(anchor="w", padx=4, pady=2)
                else:
                    w.pack_forget()

        search_var.trace_add("write", _filter)

        # 全選 / 清除
        btn_bar = ctk.CTkFrame(top, fg_color="transparent")
        btn_bar.pack(fill="x", padx=12, pady=(6, 0))
        ctk.CTkButton(btn_bar, text="全選", width=80,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"], corner_radius=8,
                       command=lambda: [v.set(True) for v in chk_vars.values()]
                       ).pack(side="left", padx=(0, 4))
        ctk.CTkButton(btn_bar, text="清除", width=80,
                       fg_color=_C["tab_bg"], text_color=_C["ink"],
                       hover_color=_C["line"], corner_radius=8,
                       command=lambda: [v.set(False) for v in chk_vars.values()]
                       ).pack(side="left")

        def apply():
            selected = {eid for eid, v in chk_vars.items() if v.get()}
            self._set_global_filter(
                emps=set() if len(selected) == len(emps) else selected
            )
            top.destroy()

        ctk.CTkButton(top, text="確定",
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=9, font=("", 13, "bold"),
                       command=apply).pack(fill="x", padx=12, pady=10)

    # --------------------------------------------------------- suggest tab
    def _build_suggest_tab(self, tab) -> None:
        _btn_outline = dict(fg_color=_C["white"], text_color=_C["ink"],
                             hover_color=_C["tab_bg"], border_width=1,
                             border_color=_C["line"], corner_radius=8)
        bar = ctk.CTkFrame(tab, fg_color=_C["white"], corner_radius=8,
                            border_width=1, border_color=_C["line"])
        bar.pack(fill="x", padx=6, pady=6)
        self.suggest_dept_btn = ctk.CTkButton(
            bar, text="部門：全部 ▼", width=_W_FILTER, **_btn_outline,
            command=self._open_suggest_dept_picker
        )
        self.suggest_emp_btn = ctk.CTkButton(
            bar, text="員工：全員 ▼", width=_W_FILTER, **_btn_outline,
            command=self._open_suggest_emp_picker
        )
        self.suggest_type_var = ctk.StringVar(value="全部")
        ctk.CTkLabel(bar, text="依打卡結果建議班別",
                      text_color=_C["muted"], font=("", 12)).pack(
            side="left", padx=(8, 6), pady=7)
        self.suggest_dept_btn.pack(side="left", padx=(0, 4), pady=5)
        self.suggest_emp_btn.pack(side="left", padx=(0, 6), pady=5)
        self.suggest_type_seg = ctk.CTkSegmentedButton(
            bar, values=["全部", "遲到早退", "加班"],
            variable=self.suggest_type_var,
            command=lambda _v: self._render_suggestions(),
            selected_color=_C["accent"],
            selected_hover_color=_C["acc_h"],
            unselected_color=_C["tab_bg"],
            unselected_hover_color=_C["line"],
            text_color=_C["ink"],
        )
        self.suggest_type_seg.pack(side="left", padx=(0, 6), pady=5)
        self.p_overtime = ctk.CTkCheckBox(bar, text="加班建議",
                                            text_color=_C["ink"],
                                            command=self._compute_suggestions)
        self.p_overtime.pack(side="left", padx=(8, 4))
        ctk.CTkButton(bar, text="重新計算", width=_W_BTN,
                       fg_color=_C["accent"], hover_color=_C["acc_h"],
                       corner_radius=8,
                       command=self._compute_suggestions).pack(
            side="left", padx=(4, 6), pady=5)
        self.suggest_count_lbl = ctk.CTkLabel(
            bar, text="共 0 筆", text_color=_C["muted"], font=("", 12)
        )
        self.suggest_count_lbl.pack(side="right", padx=8)

        body = tk.Frame(tab, bg=_C["app_bg"])
        body.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=0)
        body.rowconfigure(0, weight=1)
        self.suggest_table = _SuggestionTable(
            body, on_select=self._show_suggestion_detail,
            on_apply=self._apply_best_suggestion,
            bg=_C["app_bg"],
        )
        self.suggest_table.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        self.suggest_detail = ctk.CTkScrollableFrame(
            body, width=330, fg_color=_C["white"], corner_radius=8,
            border_width=1, border_color=_C["line"]
        )
        _bind_scrollable_frame_mousewheel(self.suggest_detail)
        self.suggest_detail.grid(row=0, column=1, sticky="nsew")
        self._suggestion_rows: list[dict] = []
        self._suggestion_by_id: dict[str, dict] = {}
        self._current_suggestion_id: str | None = None
        self._clear_suggestion_detail("請先到打卡分頁查詢，或從左側列表選擇一筆建議。")

    def _update_suggest_filter_btns(self) -> None:
        self._update_filter_buttons()

    def _suggest_row_passes_filter(self, vals: tuple) -> bool:
        emp_id = vals[0]
        dept = self._emp_filter_dept(emp_id)
        if not self._dept_passes_filter(dept):
            return False
        if self._selected_emps and emp_id not in self._selected_emps:
            return False
        return True

    def _open_suggest_dept_picker(self) -> None:
        depts = sorted({
            self._emp_filter_dept(vals[0])
            for vals, *_ in self._punch_cache
        } - {""})
        if not depts:
            messagebox.showinfo("提示", "請先在打卡分頁查詢資料。")
            return
        self._generic_picker(
            title="選擇部門",
            items=depts,
            selected=self._selected_depts,
            on_confirm=lambda s: self._set_global_filter(depts=s)
        )

    def _open_suggest_emp_picker(self) -> None:
        emps: dict[str, str] = {}
        for vals, *_ in self._punch_cache:
            emp_id, name = vals[0], vals[1]
            dept = self._emp_filter_dept(emp_id)
            if not self._dept_passes_filter(dept):
                continue
            if emp_id:
                emps[emp_id] = name
        if not emps:
            messagebox.showinfo("提示", "請先在打卡分頁查詢資料。")
            return
        items = [f"{eid}  {name}".rstrip() for eid, name in sorted(emps.items())]
        selected = {
            f"{eid}  {emps[eid]}".rstrip()
            for eid in self._selected_emps
            if eid in emps
        }
        self._generic_picker(
            title="選擇員工",
            items=items,
            selected=selected,
            on_confirm=lambda s: self._set_global_filter(
                emps={item.split()[0] for item in s}
            ),
            searchable=True,
        )

    def _compute_suggestions(self) -> None:
        if not hasattr(self, "suggest_table"):
            return
        old_rows = getattr(self, "_suggestion_by_id", {})
        rows: list[dict] = []
        filtered_cache = [
            (vals, fgs) for vals, fgs, _meta in self._punch_cache
            if self._suggest_row_passes_filter(vals)
        ]

        idx = 0
        for vals, _ in filtered_cache:
            remark = vals[8]
            if (remark and "忘打卡" not in remark
                    and any(k in remark for k in ("遲到", "早退"))):
                rows.append(self._make_suggestion_row(
                    vals, row_type="遲到早退", row_idx=idx,
                    old_rows=old_rows
                ))
                idx += 1
        if self.p_overtime.get():
            for vals, _ in filtered_cache:
                _, _, _, _, sched_in, sched_out, clock_in, clock_out, remark = vals
                if not clock_in or not clock_out or not sched_in or not sched_out:
                    continue
                if "忘打卡" in remark:
                    continue
                si_m = _punch_mins(sched_in)
                so_m = _punch_mins(sched_out)
                co_m = _punch_mins(clock_out)
                if so_m < si_m:
                    so_m += 1440
                if co_m < si_m - 120:
                    co_m += 1440
                if co_m - so_m >= 30:   # 加班 ≥ 30 分鐘才列入
                    rows.append(self._make_suggestion_row(
                        vals, row_type="加班", row_idx=idx,
                        old_rows=old_rows
                    ))
                    idx += 1
        self._suggestion_rows = rows
        self._suggestion_by_id = {row["id"]: row for row in rows}
        self._render_suggestions()

    def _make_suggestion_row(
        self, vals: tuple, row_type: str, row_idx: int,
        old_rows: dict[str, dict] | None = None,
    ) -> dict:
        emp_id, name, date, shift_code, sched_in, sched_out, \
            clock_in, clock_out, remark = vals
        cur_h = _shift_hours(shift_code) or (_calc_hours(sched_in, sched_out) if sched_in and sched_out else 8.0)
        if row_type == "加班":
            if clock_in and clock_out:
                ci_m = _punch_mins(clock_in)
                co_m = _punch_mins(clock_out)
                si_m = _punch_mins(sched_in) if sched_in else 0
                if co_m < si_m - 120:
                    co_m += 1440
                raw_h = (co_m - ci_m) / 60.0
                actual_h = raw_h - (1.0 if raw_h >= 9.0 else 0.5)
            else:
                actual_h = cur_h + 1.0
            pref_h = round(actual_h * 2) / 2   # 四捨五入到 0.5
            sugg = [s for s in _suggest_shifts(
                clock_in, clock_out, self.shift_codes, preferred_hours=pref_h, top_n=6)
                if s[5] > cur_h][:3]
        else:
            sugg = _suggest_shifts(clock_in, clock_out, self.shift_codes, preferred_hours=cur_h)
        row_id = f"{row_type}:{emp_id}:{date}:{row_idx}"
        old = (old_rows or {}).get(row_id, {})
        kind = self._default_suggestion_kind(emp_id, date, shift_code)
        best = sugg[0] if sugg else None
        shift_names = {
            code: self.shift_codes.get(code, ("", 3))[0]
            for code, *_ in sugg
        }
        return {
            "id": row_id,
            "emp_id": emp_id,
            "name": name,
            "date": date,
            "type": row_type,
            "remark": "加班建議" if row_type == "加班" else remark,
            "current_shift": shift_code or "—",
            "scheduled": self._time_range(sched_in, sched_out),
            "actual": self._time_range(clock_in, clock_out),
            "delta": self._candidate_delta_text(best),
            "best_candidate": best,
            "candidates": sugg,
            "kind": old.get("kind", kind),
            "status": old.get("status", "待處理"),
            "error": old.get("error", ""),
            "shift_names": shift_names,
            "raw": vals,
        }

    @staticmethod
    def _time_range(start: str, end: str) -> str:
        if start and end:
            return f"{start}-{end}"
        return start or end or "—"

    @staticmethod
    def _candidate_delta_text(candidate) -> str:
        if not candidate:
            return ""
        _code, _si, _so, d_in, d_out, _h = candidate
        parts: list[str] = []
        if d_in > 0:
            parts.append(f"早到{d_in}分")
        if d_out > 0:
            parts.append(f"晚下{d_out}分")
        elif d_out < 0:
            parts.append(f"早退{-d_out}分")
        return " / ".join(parts)

    def _default_suggestion_kind(self, emp_id: str, date: str, shift_code: str) -> int:
        kind = 3 if _is_rest(shift_code) else 1
        for emp in (self.schedule or []):
            if emp.emp_id != emp_id:
                continue
            existing = emp.shift_on(date)
            if existing is not None and existing.kind is not None:
                return existing.kind
            break
        return kind

    def _render_suggestions(self, select_id: str | None = None) -> None:
        if not hasattr(self, "suggest_table"):
            return
        mode = self.suggest_type_var.get() if hasattr(self, "suggest_type_var") else "全部"
        rows = list(self._suggestion_rows)
        if mode == "遲到早退":
            rows = [r for r in rows if r["type"] == "遲到早退"]
        elif mode == "加班":
            rows = [r for r in rows if r["type"] == "加班"]
        self.suggest_count_lbl.configure(text=f"共 {len(rows)} 筆")
        self.suggest_table.populate(rows)
        row_ids = {r["id"] for r in rows}
        selected = select_id if select_id in row_ids else None
        if not selected and self._current_suggestion_id in row_ids:
            selected = self._current_suggestion_id
        if not selected:
            selected = self.suggest_table.first_id()
        if selected:
            self.suggest_table.select(selected)
        else:
            self._current_suggestion_id = None
            self._clear_suggestion_detail("目前篩選條件下沒有可處理的建議")

    def _clear_suggestion_detail(self, text: str) -> None:
        for w in self.suggest_detail.winfo_children():
            w.destroy()
        ctk.CTkLabel(
            self.suggest_detail, text=text, text_color=_C["muted"],
            font=("", 13), wraplength=300, justify="center"
        ).pack(expand=True, padx=18, pady=28)

    def _show_suggestion_detail(self, row: dict | None) -> None:
        if not row:
            self._clear_suggestion_detail("目前篩選條件下沒有可處理的建議")
            return
        self._current_suggestion_id = row["id"]
        for w in self.suggest_detail.winfo_children():
            w.destroy()

        ctk.CTkLabel(
            self.suggest_detail,
            text=f"{row['name']}  {row['emp_id']}",
            text_color=_C["ink"], font=("", 15, "bold")
        ).pack(anchor="w", padx=14, pady=(14, 2))
        ctk.CTkLabel(
            self.suggest_detail,
            text=f"{row['date']}  {row['remark']}  /  {row['status']}",
            text_color=_C["muted"], font=("", 12)
        ).pack(anchor="w", padx=14, pady=(0, 10))
        if row.get("error"):
            ctk.CTkLabel(
                self.suggest_detail, text=row["error"], text_color=_C["red"],
                wraplength=310, justify="left"
            ).pack(anchor="w", padx=14, pady=(0, 8))

        summary = ctk.CTkFrame(self.suggest_detail, fg_color=_C["tab_bg"],
                               corner_radius=8)
        summary.pack(fill="x", padx=12, pady=(0, 10))
        for label, value in [
            ("原班別", row["current_shift"]),
            ("表定", row["scheduled"]),
            ("實際", row["actual"]),
            ("差異", row["delta"] or "—"),
        ]:
            line = ctk.CTkFrame(summary, fg_color="transparent")
            line.pack(fill="x", padx=10, pady=2)
            ctk.CTkLabel(line, text=label, width=52, anchor="w",
                          text_color=_C["muted"], font=("", 11)).pack(side="left")
            ctk.CTkLabel(line, text=value, anchor="w",
                          text_color=_C["ink"], font=("", 12, "bold")).pack(side="left")

        self.suggest_kind_var = tk.IntVar(value=row.get("kind", 1))
        kind_bar = ctk.CTkFrame(self.suggest_detail, fg_color=_C["white"],
                                corner_radius=8, border_width=1,
                                border_color=_C["line"])
        kind_bar.pack(fill="x", padx=12, pady=(0, 10))
        ctk.CTkLabel(kind_bar, text="排班日種類",
                      text_color=_C["ink"], font=("", 12, "bold")).pack(
            anchor="w", padx=10, pady=(8, 2))
        for label, value in [("工作日 (1)", 1), ("排休 (3)", 3)]:
            ctk.CTkRadioButton(
                kind_bar, text=label, variable=self.suggest_kind_var,
                value=value, text_color=_C["ink"], font=("", 12)
            ).pack(anchor="w", padx=10, pady=3)

        ctk.CTkLabel(
            self.suggest_detail, text="候選班別", text_color=_C["ink"],
            font=("", 13, "bold")
        ).pack(anchor="w", padx=14, pady=(2, 4))

        if not row["candidates"]:
            ctk.CTkLabel(
                self.suggest_detail, text="無合適建議",
                text_color=_C["muted"]
            ).pack(anchor="w", padx=14, pady=8)
            return

        for i, candidate in enumerate(row["candidates"]):
            code, si, so, d_in, d_out, h = candidate
            name = row["shift_names"].get(code, "")
            card = ctk.CTkFrame(
                self.suggest_detail,
                fg_color=_C["row_alt"] if i % 2 == 0 else _C["white"],
                corner_radius=8, border_width=1, border_color=_C["line"]
            )
            card.pack(fill="x", padx=12, pady=3)
            top = ctk.CTkFrame(card, fg_color="transparent")
            top.pack(fill="x", padx=10, pady=(8, 2))
            ctk.CTkLabel(
                top, text=f"{'最佳 ' if i == 0 else ''}{code}",
                text_color=_C["accent"], font=("", 13, "bold")
            ).pack(side="left")
            ctk.CTkLabel(
                top, text=f"{h:g}h", text_color=_C["muted"],
                font=("", 12)
            ).pack(side="right")
            ctk.CTkLabel(
                card, text=name or f"{si}-{so}", text_color=_C["ink"],
                anchor="w", wraplength=260
            ).pack(anchor="w", padx=10)
            ctk.CTkLabel(
                card, text=self._candidate_delta_text(candidate) or "時間吻合",
                text_color="#D07000", font=("", 11)
            ).pack(anchor="w", padx=10, pady=(0, 6))
            ctk.CTkButton(
                card, text="套用此班", height=28,
                fg_color=_C["accent"], hover_color=_C["acc_h"],
                corner_radius=6,
                command=lambda rid=row["id"], c=candidate: self._apply_suggestion_candidate(
                    rid, c, self.suggest_kind_var.get()
                )
            ).pack(fill="x", padx=10, pady=(0, 8))

    def _apply_best_suggestion(self, row_id: str) -> None:
        row = self._suggestion_by_id.get(row_id)
        if not row or not row.get("best_candidate"):
            return
        kind = getattr(self, "suggest_kind_var", tk.IntVar(value=row.get("kind", 1))).get()
        self._apply_suggestion_candidate(row_id, row["best_candidate"], kind)

    def _apply_suggestion_candidate(self, row_id: str, candidate, kind: int) -> None:
        row = self._suggestion_by_id.get(row_id)
        if not row or not candidate or not self._need_client():
            return
        code = candidate[0]
        row["kind"] = kind
        row["status"] = "套用中"
        row["error"] = ""
        self._render_suggestions(select_id=row_id)
        year, month, _ = (int(x) for x in row["date"].split("-"))
        self._set_status(f"套用中：{row['emp_id']} {row['date']} → {code}（日種類={kind}）…")

        def work():
            return self.client.set_shift(
                year, month, row["emp_id"], row["date"], code,
                kind=kind, dry_run=False,
            )

        def done(_):
            row["status"] = "已套用"
            row["error"] = ""
            self._set_status(f"已套用：{row['emp_id']} {row['date']} → {code}")
            self._render_suggestions(select_id=row_id)
            self._load_schedule()

        def failed(exc: Exception):
            row["status"] = "失敗"
            row["error"] = str(exc)
            self._set_status(f"套用失敗：{exc}")
            self._render_suggestions(select_id=row_id)
            messagebox.showerror("套用失敗", str(exc))

        self._run_async(work, done, on_error=failed)

    def _apply_suggestion(self, emp_id: str, date_str: str, code: str, kind: int = 1) -> None:
        if not self._need_client():
            return
        year, month, _ = (int(x) for x in date_str.split("-"))
        self._set_status(f"套用中：{emp_id} {date_str} → {code}（日種類={kind}）…")

        def work():
            return self.client.set_shift(
                year, month, emp_id, date_str, code,
                kind=kind,
                dry_run=False,
            )

        def done(_):
            self._set_status(f"已套用：{emp_id} {date_str} → {code}")
            self._load_schedule()

        self._run_async(work, done)


def main() -> None:
    EhrsApp().mainloop()


if __name__ == "__main__":
    main()
